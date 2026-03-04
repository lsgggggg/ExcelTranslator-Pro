#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_translator_pro.py — 产品级 Excel 翻译工具 (单文件版)
==========================================================

功能概述:
    将 Excel 文件中的文本内容从源语言翻译为目标语言，同时完整保留所有格式。
    支持 5 种翻译引擎: DeepSeek / OpenAI / Claude / Google Free / 通义千问

核心特性:
    - 方案 A (原地修改 workbook) 天然保留所有格式
    - 多 Sheet 支持, 合并单元格正确处理
    - 公式/数字/日期/编号/URL/邮箱 智能跳过
    - JSON 格式批量翻译提示词
    - 翻译缓存 + 断点续传
    - 指数退避重试 + 并发控制
    - tqdm 进度条 + logging 日志
    - 命令行 argparse 入口
    - Dry-run 模式 + 翻译统计报告

作者: Shiguang (基于原始 DeepSeek/Google 翻译脚本重构)
版本: 2.0.0
"""


# ============================================================================
# Phase 1: IMPORTS — 导入所有依赖
# ============================================================================

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import logging
import os
import re
import shutil
import signal
import sys
import textwrap
import threading
import time
import unicodedata
from abc import ABC, abstractmethod
from collections import OrderedDict, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from datetime import datetime
from enum import Enum, auto
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union

try:
    import openpyxl
    from openpyxl.cell.cell import Cell, MergedCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("错误: 缺少 openpyxl 库。请运行: pip install openpyxl")
    sys.exit(1)

try:
    from tqdm import tqdm
except ImportError:
    print("错误: 缺少 tqdm 库。请运行: pip install tqdm")
    sys.exit(1)


# ============================================================================
# Phase 1: CONSTANTS & ENUMS — 常量与枚举定义
# ============================================================================

# 版本信息
VERSION = "2.0.0"
APP_NAME = "ExcelTranslatorPro"

# 默认配置
DEFAULT_BATCH_SIZE = 30
DEFAULT_MAX_WORKERS = 3
DEFAULT_MAX_RETRIES = 3
DEFAULT_RETRY_BASE_DELAY = 1.0  # 指数退避基础延迟 (秒)
DEFAULT_RETRY_MAX_DELAY = 60.0  # 最大重试延迟 (秒)
DEFAULT_API_TIMEOUT = 120       # API 超时 (秒)
DEFAULT_SOURCE_LANG = "ru"
DEFAULT_TARGET_LANG = "zh-CN"
DEFAULT_LOG_LEVEL = "INFO"
GOOGLE_FREE_CHAR_LIMIT = 4500  # Google 免费翻译单次字符限制
MAX_CELL_LENGTH = 32767        # Excel 单元格最大字符数
CACHE_SAVE_INTERVAL = 50       # 每翻译 N 个批次后保存缓存

# 支持的翻译引擎列表
SUPPORTED_ENGINES = ["deepseek", "openai", "claude", "google_free", "qwen"]

# 各引擎环境变量名映射
ENGINE_API_KEY_ENV = {
    "deepseek": "DEEPSEEK_API_KEY",
    "openai": "OPENAI_API_KEY",
    "claude": "ANTHROPIC_API_KEY",
    "google_free": None,  # 无需 API Key
    "qwen": "DASHSCOPE_API_KEY",
}

# 各引擎默认模型
ENGINE_DEFAULT_MODELS = {
    "deepseek": "deepseek-chat",
    "openai": "gpt-4o-mini",
    "claude": "claude-sonnet-4-20250514",
    "google_free": None,
    "qwen": "qwen-plus",
}

# 各引擎默认 API Base URL
ENGINE_DEFAULT_BASE_URLS = {
    "deepseek": "https://api.deepseek.com",
    "openai": "https://api.openai.com/v1",
    "claude": None,  # Anthropic SDK 自动处理
    "google_free": None,
    "qwen": "https://dashscope.aliyuncs.com/compatible-mode/v1",
}

# 语言名称映射 (用于翻译提示词)
LANGUAGE_NAMES = {
    "zh-CN": "Simplified Chinese (简体中文)",
    "zh-TW": "Traditional Chinese (繁體中文)",
    "en": "English",
    "ru": "Russian (Русский)",
    "ja": "Japanese (日本語)",
    "ko": "Korean (한국어)",
    "fr": "French (Français)",
    "de": "German (Deutsch)",
    "es": "Spanish (Español)",
    "pt": "Portuguese (Português)",
    "ar": "Arabic (العربية)",
    "it": "Italian (Italiano)",
    "th": "Thai (ไทย)",
    "vi": "Vietnamese (Tiếng Việt)",
    "id": "Indonesian (Bahasa Indonesia)",
    "ms": "Malay (Bahasa Melayu)",
    "tr": "Turkish (Türkçe)",
    "pl": "Polish (Polski)",
    "nl": "Dutch (Nederlands)",
    "sv": "Swedish (Svenska)",
    "uk": "Ukrainian (Українська)",
    "hi": "Hindi (हिन्दी)",
}


class CellCategory(Enum):
    """单元格内容分类枚举"""
    EMPTY = auto()           # 空单元格
    FORMULA = auto()         # 公式
    NUMERIC = auto()         # 纯数字 / 货币数字
    PERCENTAGE = auto()      # 百分比
    DATE = auto()            # 日期
    CODE_LIKE = auto()       # 编号 / SKU / ID
    URL = auto()             # URL 链接
    EMAIL = auto()           # 邮箱地址
    TARGET_LANG = auto()     # 已是目标语言
    TRANSLATABLE = auto()    # 需要翻译的文本
    MERGED_SLAVE = auto()    # 合并单元格的从属单元格


# ============================================================================
# Phase 1: LOGGING 配置 — 标准化日志系统
# ============================================================================

class ColorFormatter(logging.Formatter):
    """带颜色的日志格式化器 (终端输出用)"""

    COLORS = {
        logging.DEBUG: "\033[36m",     # 青色
        logging.INFO: "\033[32m",      # 绿色
        logging.WARNING: "\033[33m",   # 黄色
        logging.ERROR: "\033[31m",     # 红色
        logging.CRITICAL: "\033[1;31m" # 加粗红色
    }
    RESET = "\033[0m"

    def format(self, record: logging.LogRecord) -> str:
        color = self.COLORS.get(record.levelno, self.RESET)
        record.levelname = f"{color}{record.levelname}{self.RESET}"
        return super().format(record)


def setup_logging(log_level: str = DEFAULT_LOG_LEVEL, log_file: Optional[str] = None) -> logging.Logger:
    """
    配置全局日志系统

    Args:
        log_level: 日志级别 (DEBUG/INFO/WARNING/ERROR/CRITICAL)
        log_file: 日志文件路径, 为 None 则自动生成

    Returns:
        配置好的 Logger 实例
    """
    logger = logging.getLogger(APP_NAME)
    logger.setLevel(getattr(logging, log_level.upper(), logging.INFO))
    logger.handlers.clear()

    # 控制台 Handler (带颜色)
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.DEBUG)
    console_fmt = ColorFormatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S"
    )
    console_handler.setFormatter(console_fmt)
    logger.addHandler(console_handler)

    # 文件 Handler
    if log_file is None:
        log_file = f"translate_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_fmt = logging.Formatter(
        fmt="%(asctime)s [%(levelname)-8s] %(funcName)s:%(lineno)d - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    file_handler.setFormatter(file_fmt)
    logger.addHandler(file_handler)

    logger.info(f"日志系统已初始化 | 级别: {log_level} | 日志文件: {log_file}")
    return logger


# 创建全局 logger (稍后在 main 中会重新配置)
logger = logging.getLogger(APP_NAME)


# ============================================================================
# Phase 1: DATA CLASSES — 数据结构定义
# ============================================================================

@dataclass
class TranslationConfig:
    """
    翻译配置数据类 — 存储所有翻译任务的配置参数

    Attributes:
        input_file: 输入 Excel 文件路径
        output_file: 输出文件路径 (可选, 默认自动生成)
        engine: 翻译引擎名称
        api_key: API 密钥 (运行时从环境变量获取)
        api_base: API 基础 URL
        model: 模型名称
        source_lang: 源语言代码
        target_lang: 目标语言代码
        batch_size: 批量翻译大小
        max_workers: 最大并发工作线程数
        max_retries: 最大重试次数
        sheets: 要翻译的 Sheet 名称列表 (为空则翻译全部)
        skip_hidden_sheets: 是否跳过隐藏的 Sheet
        cache_file: 缓存文件路径
        log_level: 日志级别
        dry_run: 是否仅做分析不实际翻译
        highlight_failures: 是否用黄色高亮标记翻译失败的单元格
        translate_comments: 是否翻译单元格批注内容
        skip_target_lang: 是否跳过已经是目标语言的文本
        generate_bilingual: 是否生成双语对照版本 Excel
        bilingual_output_file: 双语对照版本输出文件路径
    """
    input_file: str = ""
    output_file: str = ""
    engine: str = "deepseek"
    api_key: str = ""
    api_base: str = ""
    model: str = ""
    source_lang: str = DEFAULT_SOURCE_LANG
    target_lang: str = DEFAULT_TARGET_LANG
    batch_size: int = DEFAULT_BATCH_SIZE
    max_workers: int = DEFAULT_MAX_WORKERS
    max_retries: int = DEFAULT_MAX_RETRIES
    sheets: List[str] = field(default_factory=list)
    skip_hidden_sheets: bool = True
    cache_file: str = ""
    log_level: str = DEFAULT_LOG_LEVEL
    dry_run: bool = False
    highlight_failures: bool = True
    translate_comments: bool = False
    skip_target_lang: bool = True
    generate_bilingual: bool = False
    bilingual_output_file: str = ""

    def resolve_defaults(self):
        """根据引擎名称填充默认值"""
        if not self.model:
            self.model = ENGINE_DEFAULT_MODELS.get(self.engine, "")
        if not self.api_base:
            self.api_base = ENGINE_DEFAULT_BASE_URLS.get(self.engine, "") or ""
        if not self.api_key and self.engine != "google_free":
            env_var = ENGINE_API_KEY_ENV.get(self.engine, "")
            if env_var:
                self.api_key = os.environ.get(env_var, "")
        if not self.output_file:
            stem = Path(self.input_file).stem
            parent = Path(self.input_file).parent
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_file = str(parent / f"{stem}_translated_{timestamp}.xlsx")
        if not self.cache_file:
            stem = Path(self.input_file).stem
            parent = Path(self.input_file).parent
            self.cache_file = str(parent / f".{stem}_translation_cache.json")
        if self.generate_bilingual and not self.bilingual_output_file:
            stem = Path(self.input_file).stem
            parent = Path(self.input_file).parent
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.bilingual_output_file = str(parent / f"{stem}_bilingual_{timestamp}.xlsx")

    def validate(self) -> List[str]:
        """
        校验配置合法性

        Returns:
            错误信息列表, 为空表示配置有效
        """
        errors = []
        if not self.input_file:
            errors.append("必须指定输入文件路径")
        elif not os.path.exists(self.input_file):
            errors.append(f"输入文件不存在: {self.input_file}")
        elif not self.input_file.lower().endswith((".xlsx", ".xlsm")):
            errors.append("输入文件必须是 .xlsx 或 .xlsm 格式")
        if self.engine not in SUPPORTED_ENGINES:
            errors.append(f"不支持的翻译引擎: {self.engine}, 可选: {SUPPORTED_ENGINES}")
        if self.engine != "google_free" and not self.api_key and not self.dry_run:
            env_var = ENGINE_API_KEY_ENV.get(self.engine, "UNKNOWN")
            errors.append(
                f"引擎 '{self.engine}' 需要 API Key。"
                f"请设置环境变量: export {env_var}=your_api_key"
            )
        if self.batch_size < 1 or self.batch_size > 100:
            errors.append(f"batch_size 应在 1-100 之间, 当前: {self.batch_size}")
        if self.max_workers < 1 or self.max_workers > 10:
            errors.append(f"max_workers 应在 1-10 之间, 当前: {self.max_workers}")
        return errors


@dataclass
class CellInfo:
    """
    单元格信息数据类 — 存储需要翻译的单元格元数据

    Attributes:
        sheet_name: 所在 Sheet 名称
        row: 行号
        col: 列号
        original_text: 原始文本
        category: 单元格分类
        translated_text: 翻译后文本 (初始为空)
        is_translated: 是否已翻译
        error: 翻译错误信息
    """
    sheet_name: str
    row: int
    col: int
    original_text: str
    category: CellCategory
    translated_text: str = ""
    is_translated: bool = False
    error: str = ""

    @property
    def cell_ref(self) -> str:
        """获取单元格引用字符串, 如 'Sheet1!A1'"""
        return f"{self.sheet_name}!{get_column_letter(self.col)}{self.row}"

    @property
    def needs_translation(self) -> bool:
        """是否需要翻译"""
        return self.category == CellCategory.TRANSLATABLE


@dataclass
class SheetStats:
    """
    单个 Sheet 的翻译统计信息

    Attributes:
        sheet_name: Sheet 名称
        total_cells: 总单元格数 (有内容的)
        translatable: 需要翻译的单元格数
        skipped_formula: 跳过的公式单元格数
        skipped_numeric: 跳过的数字单元格数
        skipped_date: 跳过的日期单元格数
        skipped_code: 跳过的编号类单元格数
        skipped_url_email: 跳过的 URL/邮箱单元格数
        skipped_target_lang: 跳过的已是目标语言单元格数
        skipped_empty: 空单元格数
        skipped_merged: 合并从属单元格数
        translated_ok: 成功翻译数
        translated_cached: 缓存命中数
        translated_fail: 翻译失败数
        merged_regions: 合并区域数
    """
    sheet_name: str = ""
    total_cells: int = 0
    translatable: int = 0
    skipped_formula: int = 0
    skipped_numeric: int = 0
    skipped_date: int = 0
    skipped_code: int = 0
    skipped_url_email: int = 0
    skipped_target_lang: int = 0
    skipped_empty: int = 0
    skipped_merged: int = 0
    translated_ok: int = 0
    translated_cached: int = 0
    translated_fail: int = 0
    merged_regions: int = 0

    @property
    def total_skipped(self) -> int:
        return (self.skipped_formula + self.skipped_numeric + self.skipped_date
                + self.skipped_code + self.skipped_url_email + self.skipped_target_lang
                + self.skipped_empty + self.skipped_merged)


@dataclass
class TranslationReport:
    """
    翻译任务的完整统计报告

    Attributes:
        config: 翻译配置
        sheet_stats: 各 Sheet 统计
        start_time: 开始时间
        end_time: 结束时间
        total_api_calls: API 调用总次数
        total_retries: 重试总次数
        cache_hit_rate: 缓存命中率
        engine_name: 使用的翻译引擎名称
    """
    config: Optional[TranslationConfig] = None
    sheet_stats: List[SheetStats] = field(default_factory=list)
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    total_api_calls: int = 0
    total_retries: int = 0
    cache_hit_rate: float = 0.0
    engine_name: str = ""

    @property
    def total_translated(self) -> int:
        return sum(s.translated_ok for s in self.sheet_stats)

    @property
    def total_failed(self) -> int:
        return sum(s.translated_fail for s in self.sheet_stats)

    @property
    def total_cached(self) -> int:
        return sum(s.translated_cached for s in self.sheet_stats)

    @property
    def total_skipped(self) -> int:
        return sum(s.total_skipped for s in self.sheet_stats)

    @property
    def total_translatable(self) -> int:
        return sum(s.translatable for s in self.sheet_stats)

    @property
    def duration_seconds(self) -> float:
        if self.start_time and self.end_time:
            return (self.end_time - self.start_time).total_seconds()
        return 0.0

    def format_duration(self) -> str:
        """格式化持续时间为 'XmYs' 格式"""
        secs = self.duration_seconds
        if secs < 60:
            return f"{secs:.1f}s"
        mins = int(secs // 60)
        remaining_secs = secs % 60
        return f"{mins}m{remaining_secs:.0f}s"

    def print_report(self):
        """打印完整的翻译统计报告到控制台和日志"""
        sep = "=" * 68
        report_lines = [
            "",
            sep,
            f"{'翻译统计报告':^60}",
            sep,
            f"  输入文件:     {self.config.input_file if self.config else 'N/A'}",
            f"  输出文件:     {self.config.output_file if self.config else 'N/A'}",
            f"  双语对照:     {self.config.bilingual_output_file if self.config and self.config.generate_bilingual else '未生成'}",
            f"  翻译引擎:     {self.engine_name}",
            f"  源语言:       {self.config.source_lang if self.config else 'N/A'}",
            f"  目标语言:     {self.config.target_lang if self.config else 'N/A'}",
            f"  总耗时:       {self.format_duration()}",
            sep,
            f"  {'Sheet 名称':<20} {'需翻译':>8} {'成功':>8} {'缓存':>8} {'失败':>8} {'跳过':>8}",
            "-" * 68,
        ]
        for ss in self.sheet_stats:
            report_lines.append(
                f"  {ss.sheet_name:<20} {ss.translatable:>8} {ss.translated_ok:>8} "
                f"{ss.translated_cached:>8} {ss.translated_fail:>8} {ss.total_skipped:>8}"
            )
        report_lines.extend([
            "-" * 68,
            f"  {'合计':<20} {self.total_translatable:>8} {self.total_translated:>8} "
            f"{self.total_cached:>8} {self.total_failed:>8} {self.total_skipped:>8}",
            sep,
            f"  API 调用次数:   {self.total_api_calls}",
            f"  重试次数:       {self.total_retries}",
            f"  缓存命中率:     {self.cache_hit_rate:.1%}",
            sep,
        ])

        # 各 Sheet 跳过详情
        for ss in self.sheet_stats:
            if ss.total_skipped > 0:
                report_lines.append(f"  [{ss.sheet_name}] 跳过详情:")
                if ss.skipped_formula > 0:
                    report_lines.append(f"    - 公式:         {ss.skipped_formula}")
                if ss.skipped_numeric > 0:
                    report_lines.append(f"    - 数字/货币:     {ss.skipped_numeric}")
                if ss.skipped_date > 0:
                    report_lines.append(f"    - 日期:         {ss.skipped_date}")
                if ss.skipped_code > 0:
                    report_lines.append(f"    - 编号/代码:     {ss.skipped_code}")
                if ss.skipped_url_email > 0:
                    report_lines.append(f"    - URL/邮箱:     {ss.skipped_url_email}")
                if ss.skipped_target_lang > 0:
                    report_lines.append(f"    - 已是目标语言:   {ss.skipped_target_lang}")
                if ss.skipped_empty > 0:
                    report_lines.append(f"    - 空单元格:      {ss.skipped_empty}")
                if ss.skipped_merged > 0:
                    report_lines.append(f"    - 合并从属:      {ss.skipped_merged}")
        report_lines.append(sep)

        report_text = "\n".join(report_lines)
        print(report_text)
        logger.info(f"翻译统计报告:\n{report_text}")


# ============================================================================
# Phase 2: 翻译引擎基类 — TranslationEngine (ABC)
# ============================================================================

class TranslationEngine(ABC):
    """
    翻译引擎抽象基类

    所有翻译引擎 (DeepSeek / OpenAI / Claude / Google Free / 通义千问)
    都必须继承此类并实现 translate_batch 和 name 方法。

    设计原则:
        - 统一接口: 所有引擎调用方式完全一致
        - JSON 提示词: LLM 引擎使用 JSON 格式, 解析稳健
        - 容错设计: 内置重试、降级策略
    """

    def __init__(self, config: TranslationConfig):
        """
        初始化翻译引擎

        Args:
            config: 翻译配置对象
        """
        self.config = config
        self._api_call_count = 0
        self._retry_count = 0
        self._lock = threading.Lock()

    @abstractmethod
    def translate_batch(self, texts: List[str], source_lang: str,
                        target_lang: str) -> List[str]:
        """
        批量翻译文本

        Args:
            texts: 待翻译的文本列表
            source_lang: 源语言代码
            target_lang: 目标语言代码

        Returns:
            与输入等长的翻译结果列表, 翻译失败时返回原文
        """
        ...

    @property
    @abstractmethod
    def name(self) -> str:
        """翻译引擎名称标识"""
        ...

    @property
    def max_batch_size(self) -> int:
        """引擎推荐最大批量大小"""
        return 30

    @property
    def api_call_count(self) -> int:
        return self._api_call_count

    @property
    def retry_count(self) -> int:
        return self._retry_count

    def _increment_api_calls(self):
        with self._lock:
            self._api_call_count += 1

    def _increment_retries(self):
        with self._lock:
            self._retry_count += 1

    def _build_system_prompt(self, source_lang: str, target_lang: str) -> str:
        """
        构建翻译系统提示词 (适用于所有 LLM 引擎)

        Args:
            source_lang: 源语言代码
            target_lang: 目标语言代码

        Returns:
            系统提示词字符串
        """
        src_name = LANGUAGE_NAMES.get(source_lang, source_lang)
        tgt_name = LANGUAGE_NAMES.get(target_lang, target_lang)
        return (
            f"You are a professional translation engine. "
            f"Translate from {src_name} to {tgt_name}.\n\n"
            f"RULES:\n"
            f"1. Input is a JSON object where keys are IDs and values are texts to translate.\n"
            f"2. Return ONLY a valid JSON object with the same keys, values replaced by translations.\n"
            f"3. Do NOT add any explanation, markdown formatting, or extra text.\n"
            f"4. Preserve numbers, proper nouns, technical terms, and formatting.\n"
            f"5. Keep special characters (brackets, parentheses, etc.) intact.\n"
            f"6. If a value is already in the target language, keep it unchanged.\n"
            f"7. Translate naturally and idiomatically, not word-by-word.\n"
            f"8. Return PURE JSON only. No markdown code fences."
        )

    def _parse_json_response(self, response_text: str, expected_count: int,
                              original_texts: List[str]) -> List[str]:
        """
        解析 LLM 返回的 JSON 翻译结果, 带降级策略

        Args:
            response_text: LLM 返回的原始文本
            expected_count: 期望的翻译数量
            original_texts: 原始文本列表 (用于失败时回退)

        Returns:
            翻译结果列表
        """
        # 清理可能的 markdown 代码块包裹
        cleaned = response_text.strip()
        if cleaned.startswith("```"):
            # 移除 ```json 或 ``` 包裹
            lines = cleaned.split("\n")
            if lines[0].startswith("```"):
                lines = lines[1:]
            if lines and lines[-1].strip() == "```":
                lines = lines[:-1]
            cleaned = "\n".join(lines).strip()

        # 尝试直接 JSON 解析
        try:
            result = json.loads(cleaned)
            if isinstance(result, dict):
                translations = []
                for i in range(expected_count):
                    key = str(i + 1)
                    trans = result.get(key, original_texts[i])
                    if isinstance(trans, str) and trans.strip():
                        translations.append(trans.strip())
                    else:
                        translations.append(original_texts[i])
                if len(translations) == expected_count:
                    return translations
        except (json.JSONDecodeError, KeyError, IndexError):
            pass

        # 降级策略 1: 尝试查找 JSON 对象
        json_match = re.search(r'\{[^{}]*\}', cleaned, re.DOTALL)
        if json_match:
            try:
                result = json.loads(json_match.group())
                if isinstance(result, dict):
                    translations = []
                    for i in range(expected_count):
                        key = str(i + 1)
                        trans = result.get(key, original_texts[i])
                        translations.append(str(trans).strip() if trans else original_texts[i])
                    return translations
            except (json.JSONDecodeError, KeyError):
                pass

        # 降级策略 2: 尝试按行解析 (兼容编号格式)
        logger.warning("JSON 解析失败, 尝试按行解析降级策略...")
        translations = list(original_texts)
        lines = [l.strip() for l in cleaned.split("\n") if l.strip()]
        for line in lines:
            # 匹配多种编号格式: "1: xxx", "1. xxx", "#1: xxx"
            match = re.match(r'^[#"\']*(\d+)["\']?\s*[:.\-)\]]\s*(.+)', line)
            if match:
                idx = int(match.group(1)) - 1
                if 0 <= idx < expected_count:
                    translations[idx] = match.group(2).strip().strip('"\'')

        return translations

    def _exponential_backoff_retry(self, func, *args, **kwargs) -> Any:
        """
        指数退避重试包装器

        Args:
            func: 要执行的函数
            *args, **kwargs: 函数参数

        Returns:
            函数执行结果

        Raises:
            Exception: 达到最大重试次数后抛出最后一个异常
        """
        last_exception = None
        for attempt in range(1, self.config.max_retries + 1):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                last_exception = e
                self._increment_retries()
                delay = min(
                    DEFAULT_RETRY_BASE_DELAY * (2 ** (attempt - 1)),
                    DEFAULT_RETRY_MAX_DELAY
                )
                # 对于 rate limit 错误, 增加等待时间
                error_str = str(e).lower()
                if "rate" in error_str or "429" in error_str or "quota" in error_str:
                    delay = min(delay * 3, DEFAULT_RETRY_MAX_DELAY)
                    logger.warning(f"触发 API 频率限制, 等待 {delay:.1f}s 后重试...")
                else:
                    logger.warning(
                        f"API 调用失败 (尝试 {attempt}/{self.config.max_retries}): {e} | "
                        f"等待 {delay:.1f}s 后重试..."
                    )
                time.sleep(delay)

        logger.error(f"达到最大重试次数 ({self.config.max_retries}), 放弃: {last_exception}")
        raise last_exception


# ============================================================================
# Phase 2: DeepSeek 翻译引擎
# ============================================================================

class DeepSeekEngine(TranslationEngine):
    """
    DeepSeek 翻译引擎

    使用 OpenAI 兼容 SDK 调用 DeepSeek API。
    支持 JSON 格式结构化输出, 翻译质量高, 成本低。

    环境变量: DEEPSEEK_API_KEY
    默认模型: deepseek-chat
    API 地址: https://api.deepseek.com
    """

    def __init__(self, config: TranslationConfig):
        super().__init__(config)
        try:
            from openai import OpenAI
        except ImportError:
            raise ImportError("DeepSeek 引擎需要 openai 库: pip install openai")
        self.client = OpenAI(
            api_key=config.api_key,
            base_url=config.api_base or "https://api.deepseek.com",
            timeout=DEFAULT_API_TIMEOUT,
        )
        self._model = config.model or "deepseek-chat"
        logger.info(f"DeepSeek 引擎已初始化 | 模型: {self._model}")

    @property
    def name(self) -> str:
        return f"DeepSeek ({self._model})"

    @property
    def max_batch_size(self) -> int:
        return 40

    def translate_batch(self, texts: List[str], source_lang: str,
                        target_lang: str) -> List[str]:
        """
        使用 DeepSeek API 批量翻译文本

        采用 JSON 格式提示词, 使用 response_format=json_object 确保结构化输出。

        Args:
            texts: 待翻译文本列表
            source_lang: 源语言代码
            target_lang: 目标语言代码

        Returns:
            翻译结果列表
        """
        if not texts:
            return []

        numbered = {str(i + 1): t for i, t in enumerate(texts)}
        user_content = json.dumps(numbered, ensure_ascii=False, indent=None)
        system_prompt = self._build_system_prompt(source_lang, target_lang)

        def _call_api():
            self._increment_api_calls()
            response = self.client.chat.completions.create(
                model=self._model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_content},
                ],
                temperature=0.1,
                response_format={"type": "json_object"},
            )
            return response.choices[0].message.content.strip()

        try:
            result_text = self._exponential_backoff_retry(_call_api)
            return self._parse_json_response(result_text, len(texts), texts)
        except Exception as e:
            logger.error(f"DeepSeek 批量翻译最终失败: {e}")
            return list(texts)


# ============================================================================
# Phase 2: OpenAI / GPT 翻译引擎
# ============================================================================

class OpenAIEngine(TranslationEngine):
    """
    OpenAI / GPT 翻译引擎

    使用 OpenAI 官方 SDK 调用 GPT 系列模型。
    支持 JSON 模式和 function calling, 翻译质量优秀。

    环境变量: OPENAI_API_KEY
    默认模型: gpt-4o-mini
    API 地址: https://api.openai.com/v1
    """

    def __init__(self, config: TranslationConfig):
        super().__init__(config)
        try:
            from openai import OpenAI
        except ImportError:
            raise ImportError("OpenAI 引擎需要 openai 库: pip install openai")
        self.client = OpenAI(
            api_key=config.api_key,
            base_url=config.api_base or "https://api.openai.com/v1",
            timeout=DEFAULT_API_TIMEOUT,
        )
        self._model = config.model or "gpt-4o-mini"
        logger.info(f"OpenAI 引擎已初始化 | 模型: {self._model}")

    @property
    def name(self) -> str:
        return f"OpenAI ({self._model})"

    @property
    def max_batch_size(self) -> int:
        return 40

    def translate_batch(self, texts: List[str], source_lang: str,
                        target_lang: str) -> List[str]:
        """
        使用 OpenAI API 批量翻译文本

        Args:
            texts: 待翻译文本列表
            source_lang: 源语言代码
            target_lang: 目标语言代码

        Returns:
            翻译结果列表
        """
        if not texts:
            return []

        numbered = {str(i + 1): t for i, t in enumerate(texts)}
        user_content = json.dumps(numbered, ensure_ascii=False, indent=None)
        system_prompt = self._build_system_prompt(source_lang, target_lang)

        def _call_api():
            self._increment_api_calls()
            response = self.client.chat.completions.create(
                model=self._model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_content},
                ],
                temperature=0.1,
                response_format={"type": "json_object"},
            )
            return response.choices[0].message.content.strip()

        try:
            result_text = self._exponential_backoff_retry(_call_api)
            return self._parse_json_response(result_text, len(texts), texts)
        except Exception as e:
            logger.error(f"OpenAI 批量翻译最终失败: {e}")
            return list(texts)


# ============================================================================
# Phase 2: Claude (Anthropic) 翻译引擎
# ============================================================================

class ClaudeEngine(TranslationEngine):
    """
    Claude (Anthropic) 翻译引擎

    使用 Anthropic 官方 SDK 调用 Claude 系列模型。
    翻译质量极高, 尤其擅长处理复杂语境和专业术语。

    环境变量: ANTHROPIC_API_KEY
    默认模型: claude-sonnet-4-20250514
    """

    def __init__(self, config: TranslationConfig):
        super().__init__(config)
        try:
            import anthropic
        except ImportError:
            raise ImportError("Claude 引擎需要 anthropic 库: pip install anthropic")
        self._anthropic = anthropic
        self.client = anthropic.Anthropic(
            api_key=config.api_key,
            timeout=DEFAULT_API_TIMEOUT,
        )
        self._model = config.model or "claude-sonnet-4-20250514"
        logger.info(f"Claude 引擎已初始化 | 模型: {self._model}")

    @property
    def name(self) -> str:
        return f"Claude ({self._model})"

    @property
    def max_batch_size(self) -> int:
        return 35

    def translate_batch(self, texts: List[str], source_lang: str,
                        target_lang: str) -> List[str]:
        """
        使用 Anthropic Claude API 批量翻译文本

        Args:
            texts: 待翻译文本列表
            source_lang: 源语言代码
            target_lang: 目标语言代码

        Returns:
            翻译结果列表
        """
        if not texts:
            return []

        numbered = {str(i + 1): t for i, t in enumerate(texts)}
        user_content = json.dumps(numbered, ensure_ascii=False, indent=None)
        system_prompt = self._build_system_prompt(source_lang, target_lang)

        def _call_api():
            self._increment_api_calls()
            response = self.client.messages.create(
                model=self._model,
                max_tokens=4096,
                system=system_prompt,
                messages=[
                    {"role": "user", "content": user_content},
                ],
            )
            # Claude 返回 content 列表, 取第一个文本块
            return response.content[0].text.strip()

        try:
            result_text = self._exponential_backoff_retry(_call_api)
            return self._parse_json_response(result_text, len(texts), texts)
        except Exception as e:
            logger.error(f"Claude 批量翻译最终失败: {e}")
            return list(texts)


# ============================================================================
# Phase 2: Google 免费翻译引擎
# ============================================================================

class GoogleFreeEngine(TranslationEngine):
    """
    Google 免费翻译引擎

    使用 deep_translator 库调用 Google Translate (无需 API Key)。
    优点: 免费; 缺点: 单次 5000 字符限制, 频率限制严格。

    自动对超长文本进行分段处理, 内置限速保护。

    无需 API Key, 无需环境变量。
    依赖: pip install deep-translator
    """

    def __init__(self, config: TranslationConfig):
        super().__init__(config)
        try:
            from deep_translator import GoogleTranslator
        except ImportError:
            raise ImportError("Google 免费翻译引擎需要 deep-translator 库: pip install deep-translator")
        self._translator_class = GoogleTranslator
        self._rate_limit_delay = 0.15  # 每次请求间隔 (秒), 避免被封
        logger.info("Google 免费翻译引擎已初始化")

    @property
    def name(self) -> str:
        return "Google Free (deep_translator)"

    @property
    def max_batch_size(self) -> int:
        # Google 逐条翻译, batch_size 影响并发量
        return 20

    def _split_text(self, text: str, max_len: int = GOOGLE_FREE_CHAR_LIMIT) -> List[str]:
        """
        将超长文本按句子/段落边界分割为多个片段

        Args:
            text: 待分割文本
            max_len: 每个片段最大字符数

        Returns:
            分割后的文本片段列表
        """
        if len(text) <= max_len:
            return [text]

        chunks = []
        current_chunk = ""
        # 按句子分割 (中英文句号、换行)
        sentences = re.split(r'(?<=[。！？.!?\n])\s*', text)
        for sentence in sentences:
            if not sentence:
                continue
            if len(current_chunk) + len(sentence) <= max_len:
                current_chunk += sentence
            else:
                if current_chunk:
                    chunks.append(current_chunk)
                # 如果单个句子超长, 强制按字符切割
                if len(sentence) > max_len:
                    for i in range(0, len(sentence), max_len):
                        chunks.append(sentence[i:i + max_len])
                else:
                    current_chunk = sentence
        if current_chunk:
            chunks.append(current_chunk)

        return chunks if chunks else [text]

    def translate_batch(self, texts: List[str], source_lang: str,
                        target_lang: str) -> List[str]:
        """
        使用 Google 免费翻译逐条翻译文本

        对超长文本自动分段, 内置限速保护。

        Args:
            texts: 待翻译文本列表
            source_lang: 源语言代码
            target_lang: 目标语言代码

        Returns:
            翻译结果列表
        """
        results = []
        translator = self._translator_class(source=source_lang, target=target_lang)

        for text in texts:
            try:
                self._increment_api_calls()
                if len(text) > GOOGLE_FREE_CHAR_LIMIT:
                    # 超长文本分段翻译
                    chunks = self._split_text(text, GOOGLE_FREE_CHAR_LIMIT)
                    translated_parts = []
                    for chunk in chunks:
                        part = translator.translate(chunk)
                        if part:
                            translated_parts.append(part)
                        else:
                            translated_parts.append(chunk)
                        time.sleep(self._rate_limit_delay)
                    translated = "".join(translated_parts)
                else:
                    translated = translator.translate(text)

                results.append(translated if translated else text)
                time.sleep(self._rate_limit_delay)

            except Exception as e:
                logger.warning(f"Google 翻译失败: {str(e)[:100]} | 原文: {text[:50]}...")
                results.append(text)
                # 遇到错误增加等待时间
                time.sleep(self._rate_limit_delay * 5)

        return results


# ============================================================================
# Phase 2: 通义千问翻译引擎
# ============================================================================

class QwenEngine(TranslationEngine):
    """
    通义千问 (Qwen) 翻译引擎

    使用 OpenAI 兼容协议调用阿里云 DashScope API。
    支持 JSON 格式输出, 中文翻译质量优秀, 性价比高。

    环境变量: DASHSCOPE_API_KEY
    默认模型: qwen-plus
    API 地址: https://dashscope.aliyuncs.com/compatible-mode/v1
    """

    def __init__(self, config: TranslationConfig):
        super().__init__(config)
        try:
            from openai import OpenAI
        except ImportError:
            raise ImportError("通义千问引擎需要 openai 库: pip install openai")
        self.client = OpenAI(
            api_key=config.api_key,
            base_url=config.api_base or "https://dashscope.aliyuncs.com/compatible-mode/v1",
            timeout=DEFAULT_API_TIMEOUT,
        )
        self._model = config.model or "qwen-plus"
        logger.info(f"通义千问引擎已初始化 | 模型: {self._model}")

    @property
    def name(self) -> str:
        return f"Qwen ({self._model})"

    @property
    def max_batch_size(self) -> int:
        return 35

    def translate_batch(self, texts: List[str], source_lang: str,
                        target_lang: str) -> List[str]:
        """
        使用通义千问 API 批量翻译文本

        Args:
            texts: 待翻译文本列表
            source_lang: 源语言代码
            target_lang: 目标语言代码

        Returns:
            翻译结果列表
        """
        if not texts:
            return []

        numbered = {str(i + 1): t for i, t in enumerate(texts)}
        user_content = json.dumps(numbered, ensure_ascii=False, indent=None)
        system_prompt = self._build_system_prompt(source_lang, target_lang)

        def _call_api():
            self._increment_api_calls()
            response = self.client.chat.completions.create(
                model=self._model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_content},
                ],
                temperature=0.1,
                response_format={"type": "json_object"},
            )
            return response.choices[0].message.content.strip()

        try:
            result_text = self._exponential_backoff_retry(_call_api)
            return self._parse_json_response(result_text, len(texts), texts)
        except Exception as e:
            logger.error(f"通义千问批量翻译最终失败: {e}")
            return list(texts)


# ============================================================================
# Phase 2: 翻译引擎工厂 — EngineFactory
# ============================================================================

class EngineFactory:
    """
    翻译引擎工厂类

    根据配置中的引擎名称, 创建对应的翻译引擎实例。
    采用工厂模式, 使主流程与具体引擎实现解耦。
    """

    _ENGINE_MAP = {
        "deepseek": DeepSeekEngine,
        "openai": OpenAIEngine,
        "claude": ClaudeEngine,
        "google_free": GoogleFreeEngine,
        "qwen": QwenEngine,
    }

    @classmethod
    def create(cls, config: TranslationConfig) -> TranslationEngine:
        """
        根据配置创建翻译引擎

        Args:
            config: 翻译配置对象

        Returns:
            对应的翻译引擎实例

        Raises:
            ValueError: 不支持的引擎名称
            ImportError: 缺少引擎依赖库
        """
        engine_name = config.engine.lower()
        if engine_name not in cls._ENGINE_MAP:
            raise ValueError(
                f"不支持的翻译引擎: '{engine_name}'\n"
                f"可用引擎: {list(cls._ENGINE_MAP.keys())}"
            )
        engine_class = cls._ENGINE_MAP[engine_name]
        return engine_class(config)

    @classmethod
    def list_engines(cls) -> List[str]:
        """列出所有可用的翻译引擎名称"""
        return list(cls._ENGINE_MAP.keys())


# ============================================================================
# Phase 3: 单元格智能分析器 — CellAnalyzer
# ============================================================================

class CellAnalyzer:
    """
    单元格智能分析器

    负责判断每个单元格的内容类型, 决定是否需要翻译。
    包含多种检测逻辑: 公式 / 数字 / 百分比 / 日期 / 编号 / URL / 邮箱 / 语言检测。

    使用优先级:
        空值 > 公式 > 数字 > 百分比 > 日期 > URL/邮箱 > 编号 > 目标语言 > 需翻译
    """

    # ----- 预编译正则 (性能优化) -----

    # 纯数字 (含千分位/小数/货币符号/正负号)
    RE_NUMERIC = re.compile(
        r'^[+-]?[\d,]+\.?\d*\s*[¥£€$₽₹₩₿%°]*$'
        r'|^[+-]?[\d.]+,\d+\s*[¥£€$₽₹₩₿%°]*$'  # 欧洲数字格式 (逗号作小数点)
    )

    # 百分比 (含小数百分比)
    RE_PERCENTAGE = re.compile(r'^[+-]?\d[\d,.]*\s*%$')

    # 科学计数法
    RE_SCIENTIFIC = re.compile(r'^[+-]?\d+\.?\d*[eE][+-]?\d+$')

    # URL
    RE_URL = re.compile(
        r'^(https?://|ftp://|www\.)\S+$',
        re.IGNORECASE
    )

    # 邮箱
    RE_EMAIL = re.compile(
        r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    )

    # 编号 / 代码 / SKU 模式
    RE_CODE_LIKE = re.compile(
        r'^[A-Z]{2,6}[-_]?\d{2,}$'       # SKU-12345, AB123
        r'|^\d{2,}[-_][A-Z]{2,}$'         # 12345-AB
        r'|^[A-Z]\d+[A-Z]?\d*$'           # A1, B2C3
        r'|^#?\d{4,}$'                     # #12345
        r'|^\d+\.\d+\.\d+[\.\d]*$'        # 版本号 1.2.3
        r'|^[A-Z]{1,3}\d{5,}$'            # 产品编号 AB12345
        r'|^\(\d+\)$'                      # 括号编号 (123)
        r'|^[\d/]+$'                       # 纯数字/斜杠 如 2024/01/15
        r'|^\d{1,2}[./]\d{1,2}[./]\d{2,4}$'  # 日期格式 01.02.2024
    , re.IGNORECASE)

    # 日期文本模式
    RE_DATE_TEXT = re.compile(
        r'^\d{1,4}[-/.]\d{1,2}[-/.]\d{1,4}$'        # 2024-01-15 等
        r'|^\d{1,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{2,4}$'
        r'|^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{1,2},?\s+\d{2,4}$'
        r'|^\d{1,2}\s+(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)\s+\d{2,4}$'
    , re.IGNORECASE)

    # 纯标点/符号/空白
    RE_ONLY_SYMBOLS = re.compile(r'^[\s\W\d]*$', re.UNICODE)

    # CJK 字符范围 (中文)
    RE_CJK = re.compile(r'[\u4e00-\u9fff\u3400-\u4dbf]')

    # 日语假名
    RE_HIRAGANA_KATAKANA = re.compile(r'[\u3040-\u309f\u30a0-\u30ff]')

    # 韩语
    RE_HANGUL = re.compile(r'[\uac00-\ud7af\u1100-\u11ff]')

    # openpyxl 日期数字格式列表 (用于检测 number_format)
    DATE_FORMAT_CODES = {
        14, 15, 16, 17, 18, 19, 20, 21, 22,  # 内置日期格式
        27, 28, 29, 30, 31, 36, 50, 51, 52, 53, 54, 55, 56, 57, 58,  # CJK 日期
    }
    DATE_FORMAT_KEYWORDS = ['yy', 'mm', 'dd', 'hh', 'ss', 'am/pm', 'yyyy', 'mmmm']

    def __init__(self, target_lang: str = DEFAULT_TARGET_LANG,
                 skip_target_lang: bool = True):
        """
        初始化单元格分析器

        Args:
            target_lang: 目标语言代码
            skip_target_lang: 是否跳过已经是目标语言的文本
        """
        self.target_lang = target_lang
        self.skip_target_lang = skip_target_lang
        logger.debug(f"CellAnalyzer 已初始化 | 目标语言: {target_lang}")

    def analyze(self, cell: Cell, merged_slaves: Set[Tuple[int, int]]) -> CellCategory:
        """
        分析单元格内容, 返回内容分类

        Args:
            cell: openpyxl Cell 对象
            merged_slaves: 合并单元格从属坐标集合 {(row, col), ...}

        Returns:
            CellCategory 枚举值
        """
        row, col = cell.row, cell.column

        # 1. 合并单元格从属
        if (row, col) in merged_slaves:
            return CellCategory.MERGED_SLAVE

        # 2. 空单元格
        if cell.value is None:
            return CellCategory.EMPTY

        # 3. MergedCell 对象 (openpyxl 内部类型)
        if isinstance(cell, MergedCell):
            return CellCategory.MERGED_SLAVE

        # 4. 公式
        if self._is_formula(cell):
            return CellCategory.FORMULA

        # 获取文本表示
        text = self._get_text(cell)

        # 5. 空文本
        if not text or not text.strip():
            return CellCategory.EMPTY

        text_stripped = text.strip()

        # 6. 纯符号/空白
        if self.RE_ONLY_SYMBOLS.match(text_stripped):
            return CellCategory.NUMERIC

        # 7. 日期 (先检查 cell 属性, 再检查文本)
        if self._is_date(cell, text_stripped):
            return CellCategory.DATE

        # 8. 百分比
        if self.RE_PERCENTAGE.match(text_stripped):
            return CellCategory.PERCENTAGE

        # 9. 纯数字 / 科学计数法
        if self.RE_NUMERIC.match(text_stripped) or self.RE_SCIENTIFIC.match(text_stripped):
            return CellCategory.NUMERIC

        # 10. URL
        if self.RE_URL.match(text_stripped):
            return CellCategory.URL

        # 11. 邮箱
        if self.RE_EMAIL.match(text_stripped):
            return CellCategory.EMAIL

        # 12. 编号 / 代码
        if self.RE_CODE_LIKE.match(text_stripped):
            return CellCategory.CODE_LIKE

        # 13. 日期文本模式
        if self.RE_DATE_TEXT.match(text_stripped):
            return CellCategory.DATE

        # 14. 已是目标语言 (可选)
        if self.skip_target_lang and self._is_target_language(text_stripped):
            return CellCategory.TARGET_LANG

        # 15. 需要翻译
        return CellCategory.TRANSLATABLE

    def _is_formula(self, cell: Cell) -> bool:
        """
        检查单元格是否为公式

        Args:
            cell: openpyxl Cell 对象

        Returns:
            True 如果是公式
        """
        if cell.data_type == 'f':
            return True
        if isinstance(cell.value, str) and cell.value.startswith('='):
            return True
        return False

    def _get_text(self, cell: Cell) -> str:
        """
        获取单元格的文本表示

        处理特殊情况:
        - 百分比数字格式 (0.5 → "50%")
        - bool 值
        - 数字值

        Args:
            cell: openpyxl Cell 对象

        Returns:
            单元格文本内容
        """
        value = cell.value
        if value is None:
            return ""

        # 处理百分比格式的数字
        if isinstance(value, (int, float)):
            nf = cell.number_format or ""
            if '%' in nf:
                return f"{value * 100:.10g}%"
            return str(value)

        if isinstance(value, bool):
            return str(value)

        return str(value)

    def _is_date(self, cell: Cell, text: str) -> bool:
        """
        检查单元格是否为日期类型

        检查方式:
        1. cell.is_date 属性
        2. number_format 包含日期关键字
        3. 文本匹配日期模式

        Args:
            cell: openpyxl Cell 对象
            text: 单元格文本

        Returns:
            True 如果是日期
        """
        try:
            if hasattr(cell, 'is_date') and cell.is_date:
                return True
        except (AttributeError, TypeError):
            pass

        nf = (cell.number_format or "").lower()
        if nf and nf != 'general':
            for keyword in self.DATE_FORMAT_KEYWORDS:
                if keyword in nf:
                    return True

        return False

    def _is_target_language(self, text: str) -> bool:
        """
        简易检测文本是否已经是目标语言

        对于中文目标: 检查中文字符占比 > 50%
        对于日文目标: 检查假名/汉字占比
        其他语言暂不检测

        Args:
            text: 文本内容

        Returns:
            True 如果文本已经是目标语言
        """
        if not text or len(text) < 2:
            return False

        # 统计非空白字符
        non_space = re.sub(r'\s', '', text)
        if not non_space:
            return False

        total_chars = len(non_space)

        if self.target_lang.startswith("zh"):
            cjk_count = len(self.RE_CJK.findall(text))
            return (cjk_count / total_chars) > 0.5

        if self.target_lang == "ja":
            ja_count = len(self.RE_CJK.findall(text)) + len(self.RE_HIRAGANA_KATAKANA.findall(text))
            return (ja_count / total_chars) > 0.5

        if self.target_lang == "ko":
            ko_count = len(self.RE_HANGUL.findall(text))
            return (ko_count / total_chars) > 0.5

        return False

    @staticmethod
    def estimate_tokens(text: str) -> int:
        """
        粗略估算文本的 token 数量

        中文: 约 1 字 = 1.5 tokens
        英文/拉丁: 约 4 字符 = 1 token
        西里尔 (俄文): 约 2 字符 = 1 token

        Args:
            text: 文本内容

        Returns:
            估算的 token 数量
        """
        if not text:
            return 0
        cjk = len(CellAnalyzer.RE_CJK.findall(text))
        non_cjk = len(text) - cjk
        return int(cjk * 1.5 + non_cjk / 3)


# ============================================================================
# Phase 3: 翻译缓存管理器 — TranslationCache
# ============================================================================

class TranslationCache:
    """
    翻译缓存管理器 — 支持内存缓存与 JSON 文件持久化

    功能:
        - 内存缓存: 运行时即时查询, 线程安全
        - 文件持久化: 翻译中断后可从缓存恢复 (断点续传)
        - 进度记录: 记录最后翻译位置, 支持精确续传

    缓存 Key 格式: "{source_lang}:{target_lang}:{text_md5}"
        使用 MD5 哈希作为键的一部分, 避免超长文本导致 JSON 文件膨胀
    """

    def __init__(self, cache_file: Optional[str] = None):
        """
        初始化翻译缓存

        Args:
            cache_file: 缓存 JSON 文件路径, 为 None 则仅使用内存缓存
        """
        self._cache: Dict[str, str] = {}    # 内存缓存
        self._lock = threading.Lock()         # 线程锁
        self.cache_file = cache_file
        self._dirty = False                   # 是否有未保存的修改
        self._hit_count = 0
        self._miss_count = 0
        self._progress: Dict[str, Any] = {}  # 断点续传进度

        # 尝试从文件加载已有缓存
        if cache_file:
            self._load_from_file()

    def _make_key(self, text: str, source_lang: str, target_lang: str) -> str:
        """
        生成缓存 Key

        对短文本直接使用原文, 对长文本使用 MD5 哈希

        Args:
            text: 原始文本
            source_lang: 源语言
            target_lang: 目标语言

        Returns:
            缓存键字符串
        """
        if len(text) <= 200:
            return f"{source_lang}:{target_lang}:{text}"
        else:
            text_hash = hashlib.md5(text.encode('utf-8')).hexdigest()
            return f"{source_lang}:{target_lang}:md5:{text_hash}"

    def get(self, text: str, source_lang: str, target_lang: str) -> Optional[str]:
        """
        查询缓存

        Args:
            text: 原始文本
            source_lang: 源语言
            target_lang: 目标语言

        Returns:
            缓存的翻译结果, 未命中返回 None
        """
        key = self._make_key(text, source_lang, target_lang)
        with self._lock:
            result = self._cache.get(key)
            if result is not None:
                self._hit_count += 1
            else:
                self._miss_count += 1
            return result

    def put(self, text: str, translated: str, source_lang: str, target_lang: str):
        """
        写入缓存

        Args:
            text: 原始文本
            translated: 翻译结果
            source_lang: 源语言
            target_lang: 目标语言
        """
        key = self._make_key(text, source_lang, target_lang)
        with self._lock:
            self._cache[key] = translated
            self._dirty = True

    def put_batch(self, originals: List[str], translations: List[str],
                  source_lang: str, target_lang: str):
        """
        批量写入缓存

        Args:
            originals: 原始文本列表
            translations: 翻译结果列表
            source_lang: 源语言
            target_lang: 目标语言
        """
        with self._lock:
            for orig, trans in zip(originals, translations):
                if orig != trans:  # 仅缓存实际翻译了的结果
                    key = self._make_key(orig, source_lang, target_lang)
                    self._cache[key] = trans
            self._dirty = True

    def get_progress(self) -> Dict[str, Any]:
        """获取断点续传进度信息"""
        with self._lock:
            return dict(self._progress)

    def set_progress(self, sheet_name: str, row: int, col: int):
        """
        记录翻译进度 (用于断点续传)

        Args:
            sheet_name: 当前 Sheet 名称
            row: 当前行号
            col: 当前列号
        """
        with self._lock:
            self._progress = {
                "last_sheet": sheet_name,
                "last_row": row,
                "last_col": col,
                "timestamp": datetime.now().isoformat(),
            }
            self._dirty = True

    @property
    def hit_rate(self) -> float:
        """缓存命中率"""
        total = self._hit_count + self._miss_count
        return self._hit_count / total if total > 0 else 0.0

    @property
    def size(self) -> int:
        """缓存条目数"""
        return len(self._cache)

    def save(self):
        """
        将缓存持久化到 JSON 文件

        仅在有未保存的修改时才执行写入操作
        """
        if not self.cache_file or not self._dirty:
            return

        try:
            data = {
                "version": 2,
                "app": APP_NAME,
                "saved_at": datetime.now().isoformat(),
                "entries": self._cache,
                "progress": self._progress,
                "stats": {
                    "total_entries": len(self._cache),
                    "hit_count": self._hit_count,
                    "miss_count": self._miss_count,
                }
            }

            # 先写入临时文件再重命名 (原子操作, 防止写入中断导致文件损坏)
            tmp_file = self.cache_file + ".tmp"
            with open(tmp_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            shutil.move(tmp_file, self.cache_file)

            self._dirty = False
            logger.debug(f"缓存已保存: {len(self._cache)} 条 → {self.cache_file}")

        except Exception as e:
            logger.warning(f"缓存保存失败: {e}")

    def _load_from_file(self):
        """从 JSON 文件加载缓存"""
        if not self.cache_file or not os.path.exists(self.cache_file):
            return

        try:
            with open(self.cache_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if isinstance(data, dict) and "entries" in data:
                self._cache = data["entries"]
                self._progress = data.get("progress", {})
                logger.info(
                    f"从缓存文件加载 {len(self._cache)} 条翻译记录 | "
                    f"文件: {self.cache_file}"
                )
                if self._progress:
                    logger.info(
                        f"检测到断点续传进度: Sheet={self._progress.get('last_sheet')}, "
                        f"Row={self._progress.get('last_row')}, "
                        f"Col={self._progress.get('last_col')}"
                    )
            else:
                logger.warning(f"缓存文件格式不兼容, 忽略: {self.cache_file}")

        except (json.JSONDecodeError, IOError) as e:
            logger.warning(f"缓存文件加载失败, 将使用空缓存: {e}")

    def clear(self):
        """清空内存缓存和进度"""
        with self._lock:
            self._cache.clear()
            self._progress.clear()
            self._hit_count = 0
            self._miss_count = 0
            self._dirty = True


# ============================================================================
# Phase 4: Excel 处理器 — ExcelHandler (原地修改策略)
# ============================================================================

class ExcelHandler:
    """
    Excel 文件处理器 — 采用方案 A (原地修改)

    核心策略:
        直接在原始 openpyxl Workbook 对象上修改单元格的 value,
        然后另存为新文件。这样可以天然保留所有格式属性:
        字体/边框/填充/对齐/合并单元格/条件格式/数据验证/图表/图片/
        冻结窗格/隐藏行列/分组/筛选/批注/超链接/打印设置 等。

    方案 A 的优势:
        - 无需手动逐一复制 20+ 种样式属性
        - 不会因遗漏某种格式而导致输出文件丢失信息
        - 代码简洁, 维护成本低
    """

    def __init__(self, config: TranslationConfig, analyzer: CellAnalyzer):
        """
        初始化 Excel 处理器

        Args:
            config: 翻译配置
            analyzer: 单元格分析器
        """
        self.config = config
        self.analyzer = analyzer

    def load_workbook(self, filepath: str) -> openpyxl.Workbook:
        """
        加载 Excel 工作簿

        使用 data_only=False 保留公式, keep_vba=False (不处理宏)。

        Args:
            filepath: Excel 文件路径

        Returns:
            openpyxl Workbook 对象

        Raises:
            FileNotFoundError: 文件不存在
            Exception: 文件格式错误或损坏
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"文件不存在: {filepath}")

        logger.info(f"正在加载 Excel 文件: {filepath}")
        try:
            wb = openpyxl.load_workbook(
                filepath,
                data_only=False,   # 保留公式 (不计算公式结果)
                keep_links=True,   # 保留外部链接
            )
            logger.info(
                f"文件加载成功 | Sheets: {wb.sheetnames} | "
                f"总计 {len(wb.sheetnames)} 个工作表"
            )
            return wb
        except Exception as e:
            raise RuntimeError(f"无法加载 Excel 文件: {e}")

    def get_target_sheets(self, wb: openpyxl.Workbook) -> List[Worksheet]:
        """
        获取需要翻译的 Sheet 列表

        根据配置:
        - 如果指定了 sheets 参数, 则只处理指定的 Sheet
        - 否则处理全部 Sheet (可选跳过隐藏的)

        Args:
            wb: openpyxl Workbook 对象

        Returns:
            需要翻译的 Worksheet 列表
        """
        target_sheets = []

        if self.config.sheets:
            # 用户指定了 Sheet 名称
            for sheet_name in self.config.sheets:
                if sheet_name in wb.sheetnames:
                    target_sheets.append(wb[sheet_name])
                else:
                    logger.warning(f"指定的 Sheet 不存在: '{sheet_name}', 跳过")
        else:
            # 处理全部 Sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if self.config.skip_hidden_sheets and ws.sheet_state == 'hidden':
                    logger.info(f"跳过隐藏 Sheet: '{sheet_name}'")
                    continue
                target_sheets.append(ws)

        return target_sheets

    def collect_cells(self, ws: Worksheet) -> Tuple[List[CellInfo], SheetStats]:
        """
        收集指定 Sheet 中所有需要翻译的单元格

        处理流程:
        1. 识别所有合并区域, 标记从属单元格
        2. 遍历所有单元格, 用 CellAnalyzer 分析每个单元格
        3. 收集需要翻译的 CellInfo 列表
        4. 生成 SheetStats 统计信息

        Args:
            ws: openpyxl Worksheet 对象

        Returns:
            (需翻译的 CellInfo 列表, Sheet 统计信息)
        """
        stats = SheetStats(sheet_name=ws.title)

        # Step 1: 收集合并单元格的从属坐标
        merged_slaves: Set[Tuple[int, int]] = set()
        for merged_range in ws.merged_cells.ranges:
            stats.merged_regions += 1
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if row != merged_range.min_row or col != merged_range.min_col:
                        merged_slaves.add((row, col))

        # Step 2: 遍历所有单元格
        cells_to_translate: List[CellInfo] = []
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0 or max_col == 0:
            logger.info(f"  Sheet '{ws.title}' 为空, 跳过")
            return cells_to_translate, stats

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                category = self.analyzer.analyze(cell, merged_slaves)

                # 统计
                if category == CellCategory.EMPTY:
                    stats.skipped_empty += 1
                elif category == CellCategory.FORMULA:
                    stats.skipped_formula += 1
                    stats.total_cells += 1
                elif category == CellCategory.NUMERIC:
                    stats.skipped_numeric += 1
                    stats.total_cells += 1
                elif category == CellCategory.PERCENTAGE:
                    stats.skipped_numeric += 1
                    stats.total_cells += 1
                elif category == CellCategory.DATE:
                    stats.skipped_date += 1
                    stats.total_cells += 1
                elif category == CellCategory.CODE_LIKE:
                    stats.skipped_code += 1
                    stats.total_cells += 1
                elif category in (CellCategory.URL, CellCategory.EMAIL):
                    stats.skipped_url_email += 1
                    stats.total_cells += 1
                elif category == CellCategory.TARGET_LANG:
                    stats.skipped_target_lang += 1
                    stats.total_cells += 1
                elif category == CellCategory.MERGED_SLAVE:
                    stats.skipped_merged += 1
                elif category == CellCategory.TRANSLATABLE:
                    stats.total_cells += 1
                    stats.translatable += 1
                    text = self.analyzer._get_text(cell)
                    cells_to_translate.append(CellInfo(
                        sheet_name=ws.title,
                        row=row_idx,
                        col=col_idx,
                        original_text=text.strip(),
                        category=category,
                    ))

        logger.info(
            f"  Sheet '{ws.title}': 需翻译 {stats.translatable} | "
            f"跳过 {stats.total_skipped} | 合并区域 {stats.merged_regions}"
        )

        return cells_to_translate, stats

    def apply_translations(self, wb: openpyxl.Workbook,
                            translated_cells: List[CellInfo],
                            highlight_failures: bool = True):
        """
        将翻译结果写回工作簿 (原地修改)

        Args:
            wb: openpyxl Workbook 对象
            translated_cells: 已翻译的 CellInfo 列表
            highlight_failures: 是否高亮标记翻译失败的单元格
        """
        fail_fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                fill_type="solid") if highlight_failures else None

        for cell_info in translated_cells:
            ws = wb[cell_info.sheet_name]
            cell = ws.cell(row=cell_info.row, column=cell_info.col)

            if cell_info.is_translated and cell_info.translated_text:
                translated = cell_info.translated_text

                # 清理文本
                translated = translated.replace('\r\n', '\n').replace('\r', '\n')
                # 截断过长文本
                if len(translated) > MAX_CELL_LENGTH:
                    translated = translated[:MAX_CELL_LENGTH]
                    logger.warning(
                        f"  {cell_info.cell_ref}: 翻译结果超长, 已截断至 {MAX_CELL_LENGTH} 字符"
                    )

                cell.value = translated
                # 设置数字格式为文本, 防止 Excel 自动转换
                cell.number_format = '@'

            elif cell_info.error:
                # 翻译失败, 保留原文
                if highlight_failures and fail_fill:
                    cell.fill = fail_fill
                logger.debug(f"  {cell_info.cell_ref}: 翻译失败, 保留原文")

    def save_workbook(self, wb: openpyxl.Workbook, output_path: str):
        """
        保存工作簿到文件

        Args:
            wb: openpyxl Workbook 对象
            output_path: 输出文件路径
        """
        # 确保输出目录存在
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        try:
            wb.save(output_path)
            logger.info(f"文件已保存: {output_path}")
        except PermissionError:
            raise PermissionError(
                f"无法保存文件: {output_path}\n"
                f"请确保文件未被其他程序占用, 然后重试。"
            )
        except Exception as e:
            raise RuntimeError(f"保存文件失败: {e}")

    def generate_bilingual_workbook(self, input_file: str,
                                     translated_cells: List[CellInfo],
                                     output_path: str,
                                     source_lang: str = "ru",
                                     target_lang: str = "zh-CN"):
        """
        生成双语对照版本的 Excel 文件

        策略:
            在原始工作簿的基础上, 对每个包含已翻译内容的列,
            在其右侧插入一列翻译列。翻译列使用蓝色字体 + 浅蓝色背景,
            原文列保持原始格式不变, 方便用户逐一对照审核。

            处理顺序: 从最右侧列向左侧列逐一插入, 避免列偏移错乱。

        Args:
            input_file: 原始 Excel 文件路径 (从原始文件重新加载, 保持原文不变)
            translated_cells: 所有已翻译的 CellInfo 列表
            output_path: 双语对照文件输出路径
            source_lang: 源语言代码
            target_lang: 目标语言代码
        """
        logger.info("正在生成双语对照版本...")

        # 从原始文件重新加载, 保持原文不变
        wb = self.load_workbook(input_file)

        # 语言名称 (用于表头)
        src_name = LANGUAGE_NAMES.get(source_lang, source_lang).split("(")[0].strip()
        tgt_name = LANGUAGE_NAMES.get(target_lang, target_lang).split("(")[0].strip()

        # 定义翻译列样式
        trans_font = Font(color="1F4E79", name="Microsoft YaHei")  # 深蓝色字体
        trans_fill = PatternFill(
            start_color="DAEEF7", end_color="DAEEF7", fill_type="solid"  # 浅蓝色背景
        )
        header_font = Font(color="FFFFFF", name="Microsoft YaHei", bold=True, size=9)
        header_fill = PatternFill(
            start_color="2E75B6", end_color="2E75B6", fill_type="solid"  # 深蓝色表头
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        thin_border = Border(
            left=Side(style="thin", color="B4C6E7"),
            right=Side(style="thin", color="B4C6E7"),
            top=Side(style="thin", color="B4C6E7"),
            bottom=Side(style="thin", color="B4C6E7"),
        )

        # 按 Sheet 分组翻译结果
        sheet_cells: Dict[str, List[CellInfo]] = defaultdict(list)
        for cell_info in translated_cells:
            if cell_info.is_translated and cell_info.translated_text:
                sheet_cells[cell_info.sheet_name].append(cell_info)

        for sheet_name, cells in sheet_cells.items():
            if sheet_name not in wb.sheetnames:
                logger.warning(f"双语版本: Sheet '{sheet_name}' 不存在, 跳过")
                continue

            ws = wb[sheet_name]

            # 找出所有包含翻译内容的列 (去重 + 排序)
            translated_cols: Set[int] = set()
            for c in cells:
                translated_cols.add(c.col)
            sorted_cols = sorted(translated_cols, reverse=True)  # 从右向左处理

            # 建立翻译数据映射: (row, original_col) → translated_text
            trans_map: Dict[Tuple[int, int], str] = {}
            for c in cells:
                trans_map[(c.row, c.col)] = c.translated_text

            # 记录列插入偏移量: 当我们从右向左插入时, 每插入一列,
            # 其左侧的列号不受影响, 但右侧列号 +1
            # 因为从右向左处理, 所以当前列号在插入时是准确的
            col_insert_map: Dict[int, int] = {}  # 原始列号 → 插入后的翻译列号

            for orig_col in sorted_cols:
                # 在 orig_col 右侧插入一列
                insert_at = orig_col + 1
                ws.insert_cols(insert_at)
                col_insert_map[orig_col] = insert_at

                # 由于插入了列, 之前已处理的列 (在 orig_col 右侧) 的实际位置已 +1
                # 但因为我们从右向左处理, 已处理的列都在右侧,
                # 它们在 col_insert_map 中的值需要 +1
                for k in col_insert_map:
                    if k != orig_col and col_insert_map[k] > orig_col:
                        col_insert_map[k] += 1

            # 写入翻译内容
            for (row, orig_col), trans_text in trans_map.items():
                target_col = col_insert_map.get(orig_col)
                if target_col is None:
                    continue

                cell = ws.cell(row=row, column=target_col)
                # 清理翻译文本
                cleaned = trans_text.replace('\r\n', '\n').replace('\r', '\n')
                if len(cleaned) > MAX_CELL_LENGTH:
                    cleaned = cleaned[:MAX_CELL_LENGTH]
                cell.value = cleaned
                cell.font = trans_font
                cell.fill = trans_fill
                cell.border = thin_border
                cell.number_format = '@'
                # 自动换行
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

            # 为每个翻译列添加表头标记 (在第 1 行之上)
            # 找到 Sheet 的数据起始行 (第一个非空行, 通常是表头行)
            first_data_row = 1
            for row_idx in range(1, min((ws.max_row or 1) + 1, 10)):
                has_content = False
                for col_idx in range(1, (ws.max_column or 1) + 1):
                    if ws.cell(row=row_idx, column=col_idx).value is not None:
                        has_content = True
                        break
                if has_content:
                    first_data_row = row_idx
                    break

            # 设置翻译列的样式: 给整列添加浅蓝背景, 表头行加深蓝标记
            for orig_col, target_col in col_insert_map.items():
                # 表头单元格
                header_cell = ws.cell(row=first_data_row, column=target_col)
                if header_cell.value is None:
                    header_cell.value = f"▶ {tgt_name}"
                    header_cell.font = header_font
                    header_cell.fill = header_fill
                    header_cell.alignment = header_alignment
                    header_cell.border = thin_border

                # 给翻译列中所有有内容的单元格补充淡蓝背景
                for row_idx in range(first_data_row + 1, (ws.max_row or 1) + 1):
                    cell = ws.cell(row=row_idx, column=target_col)
                    if cell.value is not None:
                        if cell.fill is None or cell.fill.fgColor is None or cell.fill.patternType is None:
                            cell.fill = trans_fill
                        if cell.border is None or cell.border.left.style is None:
                            cell.border = thin_border

                # 设置翻译列宽度 (与原始列相同或稍宽)
                orig_letter = get_column_letter(orig_col)
                target_letter = get_column_letter(target_col)
                orig_width = ws.column_dimensions[orig_letter].width
                if orig_width and orig_width > 0:
                    ws.column_dimensions[target_letter].width = max(orig_width, 15)
                else:
                    ws.column_dimensions[target_letter].width = 18

        # 保存双语对照文件
        self.save_workbook(wb, output_path)
        logger.info(f"双语对照版本已生成: {output_path}")


# ============================================================================
# Phase 5: 主翻译协调器 — ExcelTranslator
# ============================================================================

class ExcelTranslator:
    """
    Excel 翻译主协调器

    负责协调整个翻译流程:
    1. 加载 Excel 文件
    2. 分析并收集需翻译的单元格
    3. 查询缓存, 过滤已翻译的内容
    4. 将未缓存的文本分批发送给翻译引擎
    5. 将翻译结果写回工作簿
    6. 保存输出文件并生成统计报告

    支持:
    - 多 Sheet 翻译
    - 并发翻译 (ThreadPoolExecutor)
    - 翻译缓存 + 断点续传
    - 实时进度显示 (tqdm)
    - 翻译失败重试与降级
    - 优雅终止 (Ctrl+C)
    """

    def __init__(self, config: TranslationConfig):
        """
        初始化翻译协调器

        Args:
            config: 翻译配置对象
        """
        self.config = config
        self.report = TranslationReport(config=config)
        self._interrupted = False

        # 初始化子组件
        self.analyzer = CellAnalyzer(
            target_lang=config.target_lang,
            skip_target_lang=config.skip_target_lang,
        )
        self.excel_handler = ExcelHandler(config, self.analyzer)
        self.cache = TranslationCache(cache_file=config.cache_file)

        # 翻译引擎 (dry_run 模式不需要)
        self.engine: Optional[TranslationEngine] = None
        if not config.dry_run:
            self.engine = EngineFactory.create(config)
            self.report.engine_name = self.engine.name

        # 注册 Ctrl+C 信号处理
        self._original_sigint = signal.getsignal(signal.SIGINT)
        signal.signal(signal.SIGINT, self._handle_interrupt)

    def _handle_interrupt(self, signum, frame):
        """
        处理 Ctrl+C 中断信号

        第一次 Ctrl+C: 设置中断标志, 优雅完成当前批次后退出
        第二次 Ctrl+C: 立即退出
        """
        if self._interrupted:
            logger.warning("\n强制退出! 缓存可能未完整保存。")
            signal.signal(signal.SIGINT, self._original_sigint)
            raise KeyboardInterrupt
        else:
            self._interrupted = True
            logger.warning(
                "\n收到中断信号! 正在完成当前批次并保存缓存... "
                "(再次按 Ctrl+C 强制退出)"
            )

    def run(self):
        """
        执行翻译主流程

        完整流程:
        1. 加载文件 → 2. 收集单元格 → 3. Dry-run/翻译 →
        4. 写入结果 → 5. 保存文件 → 6. 生成报告
        """
        self.report.start_time = datetime.now()

        try:
            # Step 1: 加载 Excel 文件
            wb = self.excel_handler.load_workbook(self.config.input_file)

            # Step 2: 获取目标 Sheet 列表
            target_sheets = self.excel_handler.get_target_sheets(wb)
            if not target_sheets:
                logger.error("没有找到需要翻译的工作表!")
                return

            logger.info(f"将翻译 {len(target_sheets)} 个工作表: "
                        f"{[ws.title for ws in target_sheets]}")

            # Step 3: 收集所有需翻译的单元格
            all_cells: List[CellInfo] = []
            for ws in target_sheets:
                cells, stats = self.excel_handler.collect_cells(ws)
                all_cells.extend(cells)
                self.report.sheet_stats.append(stats)

            total_translatable = sum(s.translatable for s in self.report.sheet_stats)
            logger.info(f"总计需翻译: {total_translatable} 个单元格")

            # Step 4: Dry-run 模式
            if self.config.dry_run:
                self._print_dry_run_report()
                return

            # Step 5: 执行翻译
            if all_cells:
                self._translate_cells(all_cells)

            # Step 6: 将翻译结果写回工作簿
            logger.info("正在将翻译结果写入工作簿...")
            self.excel_handler.apply_translations(
                wb, all_cells,
                highlight_failures=self.config.highlight_failures
            )

            # Step 7: 保存输出文件
            # 先备份原文件的副本来修改
            logger.info(f"正在保存翻译结果...")
            self.excel_handler.save_workbook(wb, self.config.output_file)

            # Step 7.5: 生成双语对照版本 (可选)
            if self.config.generate_bilingual and self.config.bilingual_output_file:
                try:
                    self.excel_handler.generate_bilingual_workbook(
                        input_file=self.config.input_file,
                        translated_cells=all_cells,
                        output_path=self.config.bilingual_output_file,
                        source_lang=self.config.source_lang,
                        target_lang=self.config.target_lang,
                    )
                except Exception as e:
                    logger.error(f"双语对照版本生成失败: {e}", exc_info=True)
                    logger.info("翻译版本已正常保存, 仅双语对照版本未生成。")

            # Step 8: 保存缓存
            self.cache.save()

        except KeyboardInterrupt:
            logger.warning("翻译被用户中断")
            self.cache.save()  # 确保中断时也保存缓存
        except Exception as e:
            logger.critical(f"翻译过程发生致命错误: {e}", exc_info=True)
            self.cache.save()
            raise
        finally:
            self.report.end_time = datetime.now()
            self.report.cache_hit_rate = self.cache.hit_rate
            if self.engine:
                self.report.total_api_calls = self.engine.api_call_count
                self.report.total_retries = self.engine.retry_count

            # 恢复原始信号处理
            signal.signal(signal.SIGINT, self._original_sigint)

            # 打印报告
            self.report.print_report()

    def _translate_cells(self, all_cells: List[CellInfo]):
        """
        执行批量翻译核心逻辑

        流程:
        1. 先查缓存, 过滤出未缓存的单元格
        2. 按 batch_size 分批
        3. 使用 ThreadPoolExecutor 并发翻译
        4. 翻译结果写入缓存
        5. 更新进度和统计

        Args:
            all_cells: 所有需翻译的 CellInfo 列表
        """
        source_lang = self.config.source_lang
        target_lang = self.config.target_lang

        # Step 1: 查缓存
        uncached_cells: List[CellInfo] = []
        for cell_info in all_cells:
            cached = self.cache.get(cell_info.original_text, source_lang, target_lang)
            if cached is not None:
                cell_info.translated_text = cached
                cell_info.is_translated = True
                # 更新对应 sheet 的 stats
                for ss in self.report.sheet_stats:
                    if ss.sheet_name == cell_info.sheet_name:
                        ss.translated_cached += 1
                        break
            else:
                uncached_cells.append(cell_info)

        cached_count = len(all_cells) - len(uncached_cells)
        if cached_count > 0:
            logger.info(f"缓存命中: {cached_count} 个单元格, 剩余待翻译: {len(uncached_cells)}")

        if not uncached_cells:
            logger.info("所有单元格均已缓存, 无需调用 API")
            return

        # Step 2: 分批
        batch_size = min(self.config.batch_size, self.engine.max_batch_size)
        batches: List[List[CellInfo]] = []
        for i in range(0, len(uncached_cells), batch_size):
            batches.append(uncached_cells[i:i + batch_size])

        logger.info(
            f"开始翻译 | 引擎: {self.engine.name} | "
            f"总计: {len(uncached_cells)} 个单元格 | "
            f"批次数: {len(batches)} | 批大小: {batch_size} | "
            f"并发: {self.config.max_workers}"
        )

        # Step 3: 并发翻译
        completed = 0
        with tqdm(total=len(uncached_cells), desc="翻译进度",
                  unit="单元格", ncols=90, colour="green") as pbar:

            # 使用线程池并发处理批次
            with ThreadPoolExecutor(max_workers=self.config.max_workers) as executor:
                # 提交所有批次
                future_to_batch = {}
                for batch_idx, batch in enumerate(batches):
                    if self._interrupted:
                        break
                    future = executor.submit(
                        self._translate_single_batch,
                        batch, source_lang, target_lang, batch_idx + 1, len(batches)
                    )
                    future_to_batch[future] = (batch_idx, batch)

                # 收集结果
                for future in as_completed(future_to_batch):
                    if self._interrupted:
                        break

                    batch_idx, batch = future_to_batch[future]
                    try:
                        future.result()  # 结果已直接写入 CellInfo
                    except Exception as e:
                        logger.error(f"批次 {batch_idx + 1} 处理异常: {e}")
                        for cell_info in batch:
                            if not cell_info.is_translated:
                                cell_info.error = str(e)
                                self._update_sheet_stats_fail(cell_info.sheet_name)

                    pbar.update(len(batch))
                    completed += len(batch)

                    # 定期保存缓存
                    if completed % (batch_size * CACHE_SAVE_INTERVAL) == 0:
                        self.cache.save()

        # 最终保存缓存
        self.cache.save()

    def _translate_single_batch(self, batch: List[CellInfo],
                                 source_lang: str, target_lang: str,
                                 batch_num: int, total_batches: int):
        """
        翻译单个批次

        Args:
            batch: 该批次的 CellInfo 列表
            source_lang: 源语言
            target_lang: 目标语言
            batch_num: 当前批次编号
            total_batches: 总批次数
        """
        texts = [c.original_text for c in batch]

        try:
            translations = self.engine.translate_batch(texts, source_lang, target_lang)

            # 校验返回数量
            if len(translations) != len(texts):
                logger.warning(
                    f"批次 {batch_num}: 返回数量不匹配 "
                    f"(期望 {len(texts)}, 实际 {len(translations)}), "
                    f"将进行补齐/截断处理"
                )
                while len(translations) < len(texts):
                    translations.append(texts[len(translations)])
                translations = translations[:len(texts)]

            # 写入结果
            for cell_info, original, translated in zip(batch, texts, translations):
                if translated and translated != original:
                    cell_info.translated_text = translated
                    cell_info.is_translated = True
                    self._update_sheet_stats_ok(cell_info.sheet_name)
                else:
                    # 翻译结果与原文相同, 可能是未翻译成功
                    cell_info.translated_text = translated or original
                    cell_info.is_translated = True
                    self._update_sheet_stats_ok(cell_info.sheet_name)

            # 批量写入缓存
            self.cache.put_batch(texts, translations, source_lang, target_lang)

            # 更新断点续传进度
            last_cell = batch[-1]
            self.cache.set_progress(last_cell.sheet_name, last_cell.row, last_cell.col)

        except Exception as e:
            logger.error(f"批次 {batch_num}/{total_batches} 翻译失败: {e}")
            for cell_info in batch:
                cell_info.error = str(e)
                self._update_sheet_stats_fail(cell_info.sheet_name)
            raise

    def _update_sheet_stats_ok(self, sheet_name: str):
        """更新指定 Sheet 的成功翻译统计"""
        for ss in self.report.sheet_stats:
            if ss.sheet_name == sheet_name:
                ss.translated_ok += 1
                break

    def _update_sheet_stats_fail(self, sheet_name: str):
        """更新指定 Sheet 的失败翻译统计"""
        for ss in self.report.sheet_stats:
            if ss.sheet_name == sheet_name:
                ss.translated_fail += 1
                break

    def _print_dry_run_report(self):
        """打印 Dry-run (仅分析) 报告"""
        sep = "=" * 68
        lines = [
            "",
            sep,
            f"{'DRY-RUN 文件分析报告':^60}",
            sep,
            f"  文件: {self.config.input_file}",
            f"  翻译方向: {self.config.source_lang} → {self.config.target_lang}",
            f"  翻译引擎: {self.config.engine}",
            sep,
        ]

        total_translatable = 0
        total_skipped = 0

        for ss in self.report.sheet_stats:
            lines.append(f"\n  📋 Sheet: \"{ss.sheet_name}\"")
            lines.append(f"     需翻译: {ss.translatable} 个单元格")
            lines.append(f"     合并区域: {ss.merged_regions} 个")
            lines.append(f"     跳过统计:")
            lines.append(f"       公式:       {ss.skipped_formula}")
            lines.append(f"       数字/货币:   {ss.skipped_numeric}")
            lines.append(f"       日期:       {ss.skipped_date}")
            lines.append(f"       编号/代码:   {ss.skipped_code}")
            lines.append(f"       URL/邮箱:   {ss.skipped_url_email}")
            lines.append(f"       已是目标语:  {ss.skipped_target_lang}")
            lines.append(f"       空单元格:    {ss.skipped_empty}")
            lines.append(f"       合并从属:    {ss.skipped_merged}")
            total_translatable += ss.translatable
            total_skipped += ss.total_skipped

        # 估算 token 数和费用
        estimated_tokens = total_translatable * 30  # 粗略估算
        lines.extend([
            "",
            sep,
            f"  📊 汇总:",
            f"     总需翻译: {total_translatable} 个单元格",
            f"     总跳过:   {total_skipped} 个单元格",
            f"     估算 Tokens: ~{estimated_tokens:,}",
            f"     预估耗时: ~{max(1, total_translatable // 50)} 分钟 (视引擎而定)",
            "",
            f"  ⚡ 提示: 移除 --dry-run 参数开始实际翻译",
            sep,
        ])

        print("\n".join(lines))


# ============================================================================
# Phase 6: CLI 命令行入口 — argparse
# ============================================================================

def build_argument_parser() -> argparse.ArgumentParser:
    """
    构建命令行参数解析器

    Returns:
        配置好的 ArgumentParser 对象
    """
    parser = argparse.ArgumentParser(
        prog="excel_translator_pro",
        description=textwrap.dedent(f"""\
            ╔══════════════════════════════════════════════╗
            ║   ExcelTranslatorPro v{VERSION}               ║
            ║   产品级 Excel 翻译工具                       ║
            ╚══════════════════════════════════════════════╝

            支持的翻译引擎:
              deepseek    - DeepSeek Chat (环境变量: DEEPSEEK_API_KEY)
              openai      - OpenAI GPT    (环境变量: OPENAI_API_KEY)
              claude      - Anthropic Claude (环境变量: ANTHROPIC_API_KEY)
              google_free - Google 翻译免费版 (无需 API Key)
              qwen        - 通义千问       (环境变量: DASHSCOPE_API_KEY)
        """),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            使用示例:
              # DeepSeek 翻译 (俄语→中文)
              python excel_translator_pro.py input.xlsx -e deepseek -s ru -t zh-CN

              # GPT-4o 翻译, 指定输出路径
              python excel_translator_pro.py input.xlsx -e openai --model gpt-4o -o output.xlsx

              # Google 免费翻译 (无需 API Key)
              python excel_translator_pro.py input.xlsx -e google_free -s ru -t zh-CN

              # 仅分析文件 (Dry-run)
              python excel_translator_pro.py input.xlsx --dry-run

              # 指定翻译特定 Sheet
              python excel_translator_pro.py input.xlsx -e deepseek --sheets "Sheet1,Sheet3"

              # Claude 翻译, 高并发
              python excel_translator_pro.py input.xlsx -e claude --workers 5 --batch 40
        """)
    )

    # ---- 必需参数 ----
    parser.add_argument(
        "input_file",
        help="输入 Excel 文件路径 (.xlsx / .xlsm)"
    )

    # ---- 翻译引擎相关 ----
    engine_group = parser.add_argument_group("翻译引擎设置")
    engine_group.add_argument(
        "-e", "--engine",
        choices=SUPPORTED_ENGINES,
        default="deepseek",
        help="翻译引擎 (默认: deepseek)"
    )
    engine_group.add_argument(
        "--api-key",
        default="",
        help="API Key (优先使用, 否则从环境变量读取)"
    )
    engine_group.add_argument(
        "--api-base",
        default="",
        help="自定义 API 基础 URL (用于代理或私有部署)"
    )
    engine_group.add_argument(
        "--model",
        default="",
        help="模型名称 (默认由引擎决定)"
    )

    # ---- 语言设置 ----
    lang_group = parser.add_argument_group("语言设置")
    lang_group.add_argument(
        "-s", "--source-lang",
        default=DEFAULT_SOURCE_LANG,
        help=f"源语言代码 (默认: {DEFAULT_SOURCE_LANG})"
    )
    lang_group.add_argument(
        "-t", "--target-lang",
        default=DEFAULT_TARGET_LANG,
        help=f"目标语言代码 (默认: {DEFAULT_TARGET_LANG})"
    )

    # ---- 输出设置 ----
    output_group = parser.add_argument_group("输出设置")
    output_group.add_argument(
        "-o", "--output",
        default="",
        help="输出文件路径 (默认: 原文件名_translated_时间戳.xlsx)"
    )
    output_group.add_argument(
        "--no-highlight",
        action="store_true",
        help="不高亮标记翻译失败的单元格"
    )
    output_group.add_argument(
        "--bilingual",
        action="store_true",
        help="同时生成双语对照版本 Excel (原文 + 翻译并排显示)"
    )

    # ---- Sheet 设置 ----
    sheet_group = parser.add_argument_group("Sheet 设置")
    sheet_group.add_argument(
        "--sheets",
        default="",
        help="指定要翻译的 Sheet 名称 (逗号分隔, 默认: 全部)"
    )
    sheet_group.add_argument(
        "--include-hidden",
        action="store_true",
        help="也翻译隐藏的 Sheet (默认跳过)"
    )

    # ---- 性能设置 ----
    perf_group = parser.add_argument_group("性能设置")
    perf_group.add_argument(
        "--batch",
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help=f"批量翻译大小 (默认: {DEFAULT_BATCH_SIZE})"
    )
    perf_group.add_argument(
        "--workers",
        type=int,
        default=DEFAULT_MAX_WORKERS,
        help=f"并发工作线程数 (默认: {DEFAULT_MAX_WORKERS})"
    )
    perf_group.add_argument(
        "--retries",
        type=int,
        default=DEFAULT_MAX_RETRIES,
        help=f"最大重试次数 (默认: {DEFAULT_MAX_RETRIES})"
    )

    # ---- 缓存与续传 ----
    cache_group = parser.add_argument_group("缓存与断点续传")
    cache_group.add_argument(
        "--cache-file",
        default="",
        help="翻译缓存文件路径 (默认自动生成)"
    )
    cache_group.add_argument(
        "--no-cache",
        action="store_true",
        help="禁用翻译缓存"
    )
    cache_group.add_argument(
        "--clear-cache",
        action="store_true",
        help="清除已有缓存后重新翻译"
    )

    # ---- 模式选项 ----
    mode_group = parser.add_argument_group("运行模式")
    mode_group.add_argument(
        "--dry-run",
        action="store_true",
        help="仅分析文件, 不执行翻译 (输出统计报告)"
    )
    mode_group.add_argument(
        "--translate-comments",
        action="store_true",
        help="同时翻译单元格批注内容"
    )
    mode_group.add_argument(
        "--no-skip-target-lang",
        action="store_true",
        help="不跳过已是目标语言的文本"
    )

    # ---- 日志设置 ----
    log_group = parser.add_argument_group("日志设置")
    log_group.add_argument(
        "--log-level",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        default=DEFAULT_LOG_LEVEL,
        help=f"日志级别 (默认: {DEFAULT_LOG_LEVEL})"
    )
    log_group.add_argument(
        "--log-file",
        default=None,
        help="日志文件路径 (默认自动生成)"
    )

    # ---- 其他 ----
    parser.add_argument(
        "-v", "--version",
        action="version",
        version=f"%(prog)s {VERSION}"
    )

    return parser


def parse_config_from_args(args: argparse.Namespace) -> TranslationConfig:
    """
    从命令行参数构建翻译配置

    Args:
        args: argparse 解析后的命名空间

    Returns:
        TranslationConfig 对象
    """
    config = TranslationConfig(
        input_file=args.input_file,
        output_file=args.output,
        engine=args.engine,
        api_key=args.api_key,
        api_base=args.api_base,
        model=args.model,
        source_lang=args.source_lang,
        target_lang=args.target_lang,
        batch_size=args.batch,
        max_workers=args.workers,
        max_retries=args.retries,
        sheets=[s.strip() for s in args.sheets.split(",") if s.strip()] if args.sheets else [],
        skip_hidden_sheets=not args.include_hidden,
        cache_file=args.cache_file if not args.no_cache else "",
        log_level=args.log_level,
        dry_run=args.dry_run,
        highlight_failures=not args.no_highlight,
        translate_comments=args.translate_comments,
        skip_target_lang=not args.no_skip_target_lang,
        generate_bilingual=args.bilingual,
    )

    # 填充默认值
    config.resolve_defaults()

    # 如果禁用缓存, 清空缓存文件路径
    if args.no_cache:
        config.cache_file = ""

    return config


# ============================================================================
# Phase 6: 交互式模式入口 — InteractiveMode
# ============================================================================

class InteractiveMode:
    """
    交互式命令行模式

    当用户不提供命令行参数时, 进入交互式引导模式,
    逐步询问用户配置参数, 降低使用门槛。
    """

    @staticmethod
    def run() -> Optional[TranslationConfig]:
        """
        运行交互式配置引导

        Returns:
            TranslationConfig 对象, 如果用户取消返回 None
        """
        print("\n" + "=" * 60)
        print(f"  ExcelTranslatorPro v{VERSION} — 交互式模式")
        print("=" * 60)

        # 输入文件
        print("\n📁 请输入 Excel 文件路径:")
        input_file = input("   > ").strip().strip('"').strip("'")
        if not input_file:
            print("❌ 未输入文件路径, 退出。")
            return None
        if not os.path.exists(input_file):
            print(f"❌ 文件不存在: {input_file}")
            return None

        # 翻译引擎
        print(f"\n🔧 选择翻译引擎 (可用: {', '.join(SUPPORTED_ENGINES)}):")
        print(f"   [回车默认: deepseek]")
        engine = input("   > ").strip().lower() or "deepseek"
        if engine not in SUPPORTED_ENGINES:
            print(f"⚠️  未知引擎 '{engine}', 使用默认 deepseek")
            engine = "deepseek"

        # 源语言
        print(f"\n🌐 源语言代码 (如 ru, en, ja, ko, fr, de):")
        print(f"   [回车默认: {DEFAULT_SOURCE_LANG}]")
        source_lang = input("   > ").strip() or DEFAULT_SOURCE_LANG

        # 目标语言
        print(f"\n🎯 目标语言代码 (如 zh-CN, en, ja):")
        print(f"   [回车默认: {DEFAULT_TARGET_LANG}]")
        target_lang = input("   > ").strip() or DEFAULT_TARGET_LANG

        # Dry-run?
        print("\n📊 是否仅分析文件 (不翻译)? [y/N]:")
        dry_run = input("   > ").strip().lower() in ('y', 'yes', '是')

        # 双语对照版本?
        generate_bilingual = False
        if not dry_run:
            print("\n📑 是否同时生成双语对照版本? (原文+翻译并排对照) [y/N]:")
            generate_bilingual = input("   > ").strip().lower() in ('y', 'yes', '是')

        # API Key 提示
        api_key = ""
        if engine != "google_free" and not dry_run:
            env_var = ENGINE_API_KEY_ENV.get(engine, "")
            existing_key = os.environ.get(env_var, "") if env_var else ""
            if existing_key:
                print(f"\n🔑 已检测到环境变量 {env_var} (将自动使用)")
            else:
                print(f"\n🔑 请输入 {engine} API Key (或设置环境变量 {env_var}):")
                api_key = input("   > ").strip()
                if not api_key:
                    print(f"⚠️  未提供 API Key, 请确保已设置环境变量 {env_var}")

        config = TranslationConfig(
            input_file=input_file,
            engine=engine,
            api_key=api_key,
            source_lang=source_lang,
            target_lang=target_lang,
            dry_run=dry_run,
            generate_bilingual=generate_bilingual,
        )
        config.resolve_defaults()

        # 确认
        print("\n" + "-" * 50)
        print("  配置确认:")
        print(f"  文件:    {config.input_file}")
        print(f"  输出:    {config.output_file}")
        print(f"  引擎:    {config.engine}")
        print(f"  方向:    {config.source_lang} → {config.target_lang}")
        print(f"  模式:    {'Dry-run (仅分析)' if config.dry_run else '正式翻译'}")
        print(f"  双语版:  {'是 ✓' if config.generate_bilingual else '否'}")
        if config.generate_bilingual:
            print(f"  双语文件: {config.bilingual_output_file}")
        print("-" * 50)
        print("\n✅ 确认开始? [Y/n]:")
        confirm = input("   > ").strip().lower()
        if confirm in ('n', 'no', '否'):
            print("❌ 已取消。")
            return None

        return config


# ============================================================================
# Phase 6: 主函数入口 — main()
# ============================================================================

def main():
    """
    程序主入口

    优先级:
    1. 如果有命令行参数, 使用 argparse 解析
    2. 如果没有参数, 进入交互式模式
    """

    # 检查是否有命令行参数 (除了脚本名)
    if len(sys.argv) > 1:
        # 命令行模式
        parser = build_argument_parser()
        args = parser.parse_args()

        # 初始化日志
        global logger
        logger = setup_logging(
            log_level=args.log_level,
            log_file=args.log_file,
        )

        # 构建配置
        config = parse_config_from_args(args)

        # 处理 --clear-cache
        if args.clear_cache and config.cache_file and os.path.exists(config.cache_file):
            os.remove(config.cache_file)
            logger.info(f"已清除缓存文件: {config.cache_file}")

    else:
        # 交互式模式
        config = InteractiveMode.run()
        if config is None:
            sys.exit(0)

        logger = setup_logging(log_level=config.log_level)

    # 校验配置
    errors = config.validate()
    if errors:
        for err in errors:
            logger.error(f"配置错误: {err}")
        sys.exit(1)

    # 打印启动信息
    logger.info(f"{'=' * 50}")
    logger.info(f"  ExcelTranslatorPro v{VERSION}")
    logger.info(f"  输入: {config.input_file}")
    logger.info(f"  输出: {config.output_file}")
    logger.info(f"  引擎: {config.engine} ({config.model or '默认模型'})")
    logger.info(f"  语言: {config.source_lang} → {config.target_lang}")
    logger.info(f"  批大小: {config.batch_size} | 并发: {config.max_workers}")
    logger.info(f"  模式: {'Dry-run' if config.dry_run else '正式翻译'}")
    if config.generate_bilingual:
        logger.info(f"  双语对照: {config.bilingual_output_file}")
    logger.info(f"{'=' * 50}")

    # 创建并运行翻译器
    translator = ExcelTranslator(config)
    translator.run()

    logger.info("程序执行完毕。")


# ============================================================================
# ENTRY POINT — 程序入口
# ============================================================================

if __name__ == "__main__":
    main()