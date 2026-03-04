#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
app.py — ExcelTranslatorPro Web Application
=============================================

基于 Flask 的 Web 界面, 将 excel_translator_pro.py 的全部功能
封装为可视化网页操作, 支持:
  - 文件上传 / 下载
  - 翻译引擎选择 & 参数配置
  - 实时翻译进度 (SSE)
  - Dry-run 文件分析
  - 双语对照版本生成
  - 翻译报告可视化

运行方式:
    pip install flask openpyxl tqdm
    python app.py

    然后浏览器打开 http://localhost:8686
"""

from __future__ import annotations

import json
import os
import queue
import shutil
import sys
import threading
import time
import uuid
import webbrowser
from datetime import datetime
from pathlib import Path

from flask import (
    Flask, render_template_string, request, jsonify,
    send_file, Response, session
)

# ─── 导入核心翻译模块 ───
# 将 excel_translator_pro.py 同目录导入
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from excel_translator_pro import (
    TranslationConfig, TranslationReport, SheetStats,
    CellAnalyzer, CellInfo, CellCategory,
    ExcelHandler, TranslationCache, EngineFactory,
    SUPPORTED_ENGINES, ENGINE_API_KEY_ENV, ENGINE_DEFAULT_MODELS,
    LANGUAGE_NAMES, DEFAULT_BATCH_SIZE, DEFAULT_MAX_WORKERS,
    DEFAULT_MAX_RETRIES, DEFAULT_SOURCE_LANG, DEFAULT_TARGET_LANG,
    VERSION, setup_logging, logger as core_logger,
)
import openpyxl
import logging

# ============================================================================
# Flask App 初始化
# ============================================================================

app = Flask(__name__)
app.secret_key = os.urandom(24)

# 工作目录
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
OUTPUT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# 全局任务状态存储
tasks: dict = {}  # task_id -> TaskState

# ============================================================================
# 任务状态管理
# ============================================================================

class TaskState:
    """翻译任务的运行时状态"""
    def __init__(self, task_id: str):
        self.task_id = task_id
        self.status = "pending"  # pending / analyzing / translating / done / error
        self.progress = 0        # 0-100
        self.total_cells = 0
        self.translated_cells = 0
        self.current_sheet = ""
        self.message = ""
        self.error = ""
        self.report: dict = {}
        self.output_file = ""
        self.bilingual_file = ""
        self.input_filename = ""
        self.event_queue = queue.Queue()
        self.lock = threading.Lock()

    def send_event(self, event_type: str, data: dict):
        """推送 SSE 事件"""
        self.event_queue.put({"type": event_type, "data": data})

    def to_dict(self):
        with self.lock:
            return {
                "task_id": self.task_id,
                "status": self.status,
                "progress": self.progress,
                "total_cells": self.total_cells,
                "translated_cells": self.translated_cells,
                "current_sheet": self.current_sheet,
                "message": self.message,
                "error": self.error,
                "report": self.report,
                "output_file": self.output_file,
                "bilingual_file": self.bilingual_file,
                "input_filename": self.input_filename,
            }


# ============================================================================
# 后端翻译执行线程
# ============================================================================

def run_translation_task(task_id: str, config: TranslationConfig,
                         input_path: str, task_state: TaskState):
    """
    在后台线程中运行翻译任务, 通过 task_state 推送进度事件
    """
    log = logging.getLogger("WebTranslator")

    try:
        task_state.status = "analyzing"
        task_state.message = "正在分析 Excel 文件..."
        task_state.send_event("status", {"status": "analyzing", "message": "正在分析文件结构..."})

        # ─── 初始化组件 ───
        analyzer = CellAnalyzer(
            target_lang=config.target_lang,
            skip_target_lang=config.skip_target_lang,
        )
        excel_handler = ExcelHandler(config, analyzer)
        cache = TranslationCache(cache_file=config.cache_file if config.cache_file else None)

        # ─── 加载工作簿 ───
        wb = excel_handler.load_workbook(input_path)
        target_sheets = excel_handler.get_target_sheets(wb)

        if not target_sheets:
            task_state.status = "error"
            task_state.error = "没有找到需要翻译的工作表"
            task_state.send_event("error", {"error": task_state.error})
            return

        # ─── 收集单元格 ───
        all_cells = []
        sheet_stats_list = []
        for ws in target_sheets:
            cells, stats = excel_handler.collect_cells(ws)
            all_cells.extend(cells)
            sheet_stats_list.append(stats)

        total_translatable = sum(s.translatable for s in sheet_stats_list)
        task_state.total_cells = total_translatable

        task_state.send_event("analysis", {
            "total_translatable": total_translatable,
            "sheets": [
                {
                    "name": s.sheet_name,
                    "translatable": s.translatable,
                    "skipped_formula": s.skipped_formula,
                    "skipped_numeric": s.skipped_numeric,
                    "skipped_date": s.skipped_date,
                    "skipped_code": s.skipped_code,
                    "skipped_url_email": s.skipped_url_email,
                    "skipped_target_lang": s.skipped_target_lang,
                    "skipped_empty": s.skipped_empty,
                    "skipped_merged": s.skipped_merged,
                    "merged_regions": s.merged_regions,
                    "total_skipped": s.total_skipped,
                }
                for s in sheet_stats_list
            ]
        })

        # ─── Dry-run 模式 ───
        if config.dry_run:
            task_state.status = "done"
            task_state.message = "分析完成 (Dry-run 模式)"
            task_state.progress = 100
            task_state.report = _build_report_dict(sheet_stats_list, config, 0, 0, 0.0, "Dry-run")
            task_state.send_event("done", {"message": "分析完成", "report": task_state.report})
            return

        # ─── 创建翻译引擎 ───
        task_state.status = "translating"
        task_state.message = "正在初始化翻译引擎..."
        task_state.send_event("status", {"status": "translating", "message": "初始化翻译引擎..."})

        engine = EngineFactory.create(config)

        # ─── 查缓存 ───
        uncached_cells = []
        for cell_info in all_cells:
            cached = cache.get(cell_info.original_text, config.source_lang, config.target_lang)
            if cached is not None:
                cell_info.translated_text = cached
                cell_info.is_translated = True
                for ss in sheet_stats_list:
                    if ss.sheet_name == cell_info.sheet_name:
                        ss.translated_cached += 1
                        break
            else:
                uncached_cells.append(cell_info)

        cached_count = len(all_cells) - len(uncached_cells)
        if cached_count > 0:
            task_state.send_event("cache", {"cached": cached_count, "remaining": len(uncached_cells)})

        # ─── 批量翻译 ───
        if uncached_cells:
            batch_size = min(config.batch_size, engine.max_batch_size)
            batches = [uncached_cells[i:i + batch_size] for i in range(0, len(uncached_cells), batch_size)]
            completed = 0

            task_state.send_event("status", {
                "status": "translating",
                "message": f"开始翻译 {len(uncached_cells)} 个单元格 ({len(batches)} 批次)"
            })

            for batch_idx, batch in enumerate(batches):
                texts = [c.original_text for c in batch]
                current_sheet = batch[0].sheet_name if batch else ""
                task_state.current_sheet = current_sheet

                try:
                    translations = engine.translate_batch(texts, config.source_lang, config.target_lang)

                    # 长度修正
                    while len(translations) < len(texts):
                        translations.append(texts[len(translations)])
                    translations = translations[:len(texts)]

                    for cell_info, original, translated in zip(batch, texts, translations):
                        cell_info.translated_text = translated or original
                        cell_info.is_translated = True
                        for ss in sheet_stats_list:
                            if ss.sheet_name == cell_info.sheet_name:
                                ss.translated_ok += 1
                                break

                    cache.put_batch(texts, translations, config.source_lang, config.target_lang)

                except Exception as e:
                    log.error(f"批次 {batch_idx + 1} 翻译失败: {e}")
                    for cell_info in batch:
                        cell_info.error = str(e)
                        for ss in sheet_stats_list:
                            if ss.sheet_name == cell_info.sheet_name:
                                ss.translated_fail += 1
                                break

                completed += len(batch)
                progress = int(completed / len(uncached_cells) * 100) if uncached_cells else 100
                task_state.translated_cells = completed + cached_count
                task_state.progress = progress

                task_state.send_event("progress", {
                    "progress": progress,
                    "completed": completed + cached_count,
                    "total": total_translatable,
                    "batch": batch_idx + 1,
                    "total_batches": len(batches),
                    "sheet": current_sheet,
                })

            cache.save()
        else:
            task_state.progress = 100

        # ─── 写回工作簿 ───
        task_state.message = "正在写入翻译结果..."
        task_state.send_event("status", {"status": "saving", "message": "写入翻译结果..."})
        excel_handler.apply_translations(wb, all_cells, highlight_failures=config.highlight_failures)

        # ─── 保存文件 ───
        excel_handler.save_workbook(wb, config.output_file)
        task_state.output_file = config.output_file

        # ─── 双语对照 ───
        if config.generate_bilingual and config.bilingual_output_file:
            task_state.message = "正在生成双语对照版本..."
            task_state.send_event("status", {"status": "bilingual", "message": "生成双语对照版本..."})
            try:
                excel_handler.generate_bilingual_workbook(
                    input_file=input_path,
                    translated_cells=all_cells,
                    output_path=config.bilingual_output_file,
                    source_lang=config.source_lang,
                    target_lang=config.target_lang,
                )
                task_state.bilingual_file = config.bilingual_output_file
            except Exception as e:
                log.error(f"双语对照生成失败: {e}")

        # ─── 完成 ───
        task_state.status = "done"
        task_state.progress = 100
        task_state.message = "翻译完成!"

        report = _build_report_dict(
            sheet_stats_list, config,
            engine.api_call_count, engine.retry_count,
            cache.hit_rate, engine.name
        )
        task_state.report = report
        task_state.send_event("done", {"message": "翻译完成!", "report": report})

    except Exception as e:
        log.error(f"翻译任务异常: {e}", exc_info=True)
        task_state.status = "error"
        task_state.error = str(e)
        task_state.send_event("error", {"error": str(e)})


def _build_report_dict(sheet_stats_list, config, api_calls, retries, hit_rate, engine_name):
    """构建报告字典"""
    return {
        "engine": engine_name,
        "source_lang": config.source_lang,
        "target_lang": config.target_lang,
        "api_calls": api_calls,
        "retries": retries,
        "cache_hit_rate": round(hit_rate * 100, 1),
        "sheets": [
            {
                "name": s.sheet_name,
                "translatable": s.translatable,
                "translated_ok": s.translated_ok,
                "translated_cached": s.translated_cached,
                "translated_fail": s.translated_fail,
                "total_skipped": s.total_skipped,
                "skipped_formula": s.skipped_formula,
                "skipped_numeric": s.skipped_numeric,
                "skipped_date": s.skipped_date,
                "skipped_code": s.skipped_code,
                "skipped_url_email": s.skipped_url_email,
                "skipped_target_lang": s.skipped_target_lang,
                "skipped_empty": s.skipped_empty,
                "skipped_merged": s.skipped_merged,
            }
            for s in sheet_stats_list
        ],
        "total_translatable": sum(s.translatable for s in sheet_stats_list),
        "total_ok": sum(s.translated_ok for s in sheet_stats_list),
        "total_cached": sum(s.translated_cached for s in sheet_stats_list),
        "total_fail": sum(s.translated_fail for s in sheet_stats_list),
        "total_skipped": sum(s.total_skipped for s in sheet_stats_list),
    }


# ============================================================================
# Flask 路由
# ============================================================================

@app.route("/")
def index():
    """主页"""
    return render_template_string(HTML_TEMPLATE)


@app.route("/api/upload", methods=["POST"])
def upload_file():
    """上传 Excel 文件并返回 Sheet 信息"""
    if "file" not in request.files:
        return jsonify({"error": "未选择文件"}), 400

    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "文件名为空"}), 400

    ext = Path(f.filename).suffix.lower()
    if ext not in (".xlsx", ".xlsm"):
        return jsonify({"error": "仅支持 .xlsx / .xlsm 格式"}), 400

    # 保存文件
    file_id = str(uuid.uuid4())[:8]
    safe_name = f"{file_id}_{f.filename}"
    filepath = os.path.join(UPLOAD_FOLDER, safe_name)
    f.save(filepath)

    # 解析 Sheet 信息
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        sheets_info = []
        for name in wb.sheetnames:
            ws = wb[name]
            is_hidden = ws.sheet_state == "hidden"
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            sheets_info.append({
                "name": name,
                "hidden": is_hidden,
                "rows": max_row,
                "cols": max_col,
            })
        wb.close()
    except Exception as e:
        return jsonify({"error": f"文件解析失败: {str(e)}"}), 400

    return jsonify({
        "file_id": file_id,
        "filename": f.filename,
        "filepath": filepath,
        "sheets": sheets_info,
    })


@app.route("/api/translate", methods=["POST"])
def start_translation():
    """启动翻译任务"""
    data = request.json
    if not data:
        return jsonify({"error": "无效请求"}), 400

    filepath = data.get("filepath", "")
    if not filepath or not os.path.exists(filepath):
        return jsonify({"error": "文件不存在, 请重新上传"}), 400

    # 构建配置
    task_id = str(uuid.uuid4())[:12]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    original_name = Path(data.get("filename", "output")).stem

    output_path = os.path.join(OUTPUT_FOLDER, f"{original_name}_translated_{timestamp}.xlsx")
    bilingual_path = os.path.join(OUTPUT_FOLDER, f"{original_name}_bilingual_{timestamp}.xlsx")
    cache_path = os.path.join(UPLOAD_FOLDER, f".cache_{original_name}.json")

    engine = data.get("engine", "deepseek")
    api_key = data.get("api_key", "")

    # 如果没提供 API key, 尝试从环境变量获取
    if not api_key and engine != "google_free":
        env_var = ENGINE_API_KEY_ENV.get(engine, "")
        if env_var:
            api_key = os.environ.get(env_var, "")

    config = TranslationConfig(
        input_file=filepath,
        output_file=output_path,
        engine=engine,
        api_key=api_key,
        api_base=data.get("api_base", ""),
        model=data.get("model", ""),
        source_lang=data.get("source_lang", DEFAULT_SOURCE_LANG),
        target_lang=data.get("target_lang", DEFAULT_TARGET_LANG),
        batch_size=int(data.get("batch_size", DEFAULT_BATCH_SIZE)),
        max_workers=int(data.get("max_workers", DEFAULT_MAX_WORKERS)),
        max_retries=int(data.get("max_retries", DEFAULT_MAX_RETRIES)),
        sheets=[s.strip() for s in data.get("sheets", "").split(",") if s.strip()] if data.get("sheets") else [],
        skip_hidden_sheets=data.get("skip_hidden", True),
        cache_file=cache_path if not data.get("no_cache", False) else "",
        dry_run=data.get("dry_run", False),
        highlight_failures=data.get("highlight_failures", True),
        translate_comments=data.get("translate_comments", False),
        skip_target_lang=data.get("skip_target_lang", True),
        generate_bilingual=data.get("bilingual", False),
        bilingual_output_file=bilingual_path if data.get("bilingual", False) else "",
    )
    config.resolve_defaults()

    # 验证
    errors = config.validate()
    if errors:
        return jsonify({"error": " | ".join(errors)}), 400

    # 创建任务
    task_state = TaskState(task_id)
    task_state.input_filename = data.get("filename", "")
    tasks[task_id] = task_state

    # 启动后台线程
    t = threading.Thread(
        target=run_translation_task,
        args=(task_id, config, filepath, task_state),
        daemon=True,
    )
    t.start()

    return jsonify({"task_id": task_id})


@app.route("/api/task/<task_id>/status")
def task_status(task_id):
    """查询任务状态"""
    ts = tasks.get(task_id)
    if not ts:
        return jsonify({"error": "任务不存在"}), 404
    return jsonify(ts.to_dict())


@app.route("/api/task/<task_id>/stream")
def task_stream(task_id):
    """SSE 事件流 — 实时推送翻译进度"""
    ts = tasks.get(task_id)
    if not ts:
        return jsonify({"error": "任务不存在"}), 404

    def event_generator():
        while True:
            try:
                event = ts.event_queue.get(timeout=30)
                yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
                if event.get("type") in ("done", "error"):
                    break
            except queue.Empty:
                yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"
            except GeneratorExit:
                break

    return Response(
        event_generator(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        }
    )


@app.route("/api/download/<task_id>/<file_type>")
def download_file(task_id, file_type):
    """下载翻译结果文件"""
    ts = tasks.get(task_id)
    if not ts:
        return jsonify({"error": "任务不存在"}), 404

    if file_type == "translated":
        fpath = ts.output_file
        fname = f"{Path(ts.input_filename).stem}_translated.xlsx"
    elif file_type == "bilingual":
        fpath = ts.bilingual_file
        fname = f"{Path(ts.input_filename).stem}_bilingual.xlsx"
    else:
        return jsonify({"error": "无效文件类型"}), 400

    if not fpath or not os.path.exists(fpath):
        return jsonify({"error": "文件不存在"}), 404

    return send_file(fpath, as_attachment=True, download_name=fname)


@app.route("/api/engines")
def get_engines():
    """返回可用翻译引擎与语言列表"""
    env_status = {}
    for eng, env_var in ENGINE_API_KEY_ENV.items():
        if env_var:
            env_status[eng] = bool(os.environ.get(env_var, ""))
        else:
            env_status[eng] = True  # google_free 不需要 key

    return jsonify({
        "engines": SUPPORTED_ENGINES,
        "default_models": ENGINE_DEFAULT_MODELS,
        "env_status": env_status,
        "languages": LANGUAGE_NAMES,
    })


# ============================================================================
# HTML 模板 — 完整的单页应用
# ============================================================================

HTML_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ExcelTranslator Pro — 智能 Excel 翻译平台</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;1,9..40,400&family=JetBrains+Mono:wght@400;500&family=Noto+Sans+SC:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
/* ══════════════════════════════════════════════════════════════
   CSS VARIABLES & RESET
   ══════════════════════════════════════════════════════════════ */
:root {
  --bg-primary: #FAFBFD;
  --bg-card: #FFFFFF;
  --bg-elevated: #F4F6FA;
  --bg-subtle: #EEF1F8;
  --bg-accent-soft: #EBF2FF;
  --border-light: #E2E6EF;
  --border-medium: #D0D5E0;
  --text-primary: #1A1D26;
  --text-secondary: #5A6070;
  --text-tertiary: #8B92A5;
  --text-inverse: #FFFFFF;
  --accent: #2563EB;
  --accent-hover: #1D4FD7;
  --accent-soft: #DBEAFE;
  --accent-text: #1E40AF;
  --success: #059669;
  --success-soft: #D1FAE5;
  --warning: #D97706;
  --warning-soft: #FEF3C7;
  --error: #DC2626;
  --error-soft: #FEE2E2;
  --info: #7C3AED;
  --info-soft: #EDE9FE;
  --shadow-sm: 0 1px 2px rgba(0,0,0,0.04), 0 1px 3px rgba(0,0,0,0.06);
  --shadow-md: 0 4px 6px -1px rgba(0,0,0,0.06), 0 2px 4px -2px rgba(0,0,0,0.05);
  --shadow-lg: 0 10px 25px -5px rgba(0,0,0,0.08), 0 8px 10px -6px rgba(0,0,0,0.04);
  --shadow-xl: 0 20px 50px -12px rgba(0,0,0,0.12);
  --radius-sm: 8px;
  --radius-md: 12px;
  --radius-lg: 16px;
  --radius-xl: 20px;
  --font-sans: 'DM Sans', 'Noto Sans SC', -apple-system, BlinkMacSystemFont, sans-serif;
  --font-mono: 'JetBrains Mono', 'SF Mono', 'Fira Code', monospace;
  --transition: 0.2s cubic-bezier(0.4, 0, 0.2, 1);
}

*, *::before, *::after { margin:0; padding:0; box-sizing:border-box; }

html { scroll-behavior: smooth; }

body {
  font-family: var(--font-sans);
  background: var(--bg-primary);
  color: var(--text-primary);
  line-height: 1.6;
  min-height: 100vh;
  -webkit-font-smoothing: antialiased;
  -moz-osx-font-smoothing: grayscale;
}

/* ══════════════════════════════════════════════════════════════
   GEOMETRIC BACKGROUND PATTERN
   ══════════════════════════════════════════════════════════════ */
body::before {
  content: '';
  position: fixed;
  inset: 0;
  background:
    radial-gradient(ellipse 80% 50% at 20% -10%, rgba(37,99,235,0.06) 0%, transparent 60%),
    radial-gradient(ellipse 60% 40% at 80% 110%, rgba(124,58,237,0.04) 0%, transparent 50%),
    linear-gradient(180deg, #FAFBFD 0%, #F0F2F8 100%);
  pointer-events: none;
  z-index: -1;
}

/* ══════════════════════════════════════════════════════════════
   HEADER / NAVIGATION
   ══════════════════════════════════════════════════════════════ */
.site-header {
  position: sticky;
  top: 0;
  z-index: 100;
  background: rgba(250,251,253,0.85);
  backdrop-filter: blur(20px) saturate(1.4);
  -webkit-backdrop-filter: blur(20px) saturate(1.4);
  border-bottom: 1px solid var(--border-light);
}
.header-inner {
  max-width: 1200px;
  margin: 0 auto;
  padding: 0 32px;
  height: 64px;
  display: flex;
  align-items: center;
  justify-content: space-between;
}
.logo {
  display: flex;
  align-items: center;
  gap: 12px;
  text-decoration: none;
}
.logo-icon {
  width: 36px; height: 36px;
  background: linear-gradient(135deg, var(--accent), #7C3AED);
  border-radius: 10px;
  display: flex;
  align-items: center;
  justify-content: center;
  color: white;
  font-size: 18px;
  font-weight: 700;
  letter-spacing: -0.5px;
  box-shadow: 0 2px 8px rgba(37,99,235,0.3);
}
.logo-text {
  font-size: 18px;
  font-weight: 700;
  color: var(--text-primary);
  letter-spacing: -0.3px;
}
.logo-text span { color: var(--accent); }
.header-badge {
  font-size: 11px;
  font-weight: 600;
  padding: 3px 8px;
  border-radius: 6px;
  background: var(--accent-soft);
  color: var(--accent-text);
  letter-spacing: 0.3px;
}

/* ══════════════════════════════════════════════════════════════
   HERO SECTION
   ══════════════════════════════════════════════════════════════ */
.hero {
  max-width: 1200px;
  margin: 0 auto;
  padding: 48px 32px 24px;
  text-align: center;
}
.hero h1 {
  font-size: 40px;
  font-weight: 700;
  letter-spacing: -1px;
  line-height: 1.15;
  margin-bottom: 14px;
  background: linear-gradient(135deg, var(--text-primary) 30%, var(--accent) 100%);
  -webkit-background-clip: text;
  -webkit-text-fill-color: transparent;
  background-clip: text;
}
.hero p {
  font-size: 17px;
  color: var(--text-secondary);
  max-width: 580px;
  margin: 0 auto;
  line-height: 1.7;
}
.engine-pills {
  display: flex;
  justify-content: center;
  gap: 8px;
  margin-top: 20px;
  flex-wrap: wrap;
}
.engine-pill {
  font-size: 12px;
  font-weight: 500;
  padding: 5px 14px;
  border-radius: 20px;
  background: var(--bg-card);
  border: 1px solid var(--border-light);
  color: var(--text-secondary);
  transition: var(--transition);
}
.engine-pill:hover {
  border-color: var(--accent);
  color: var(--accent);
  background: var(--accent-soft);
}

/* ══════════════════════════════════════════════════════════════
   MAIN CONTAINER
   ══════════════════════════════════════════════════════════════ */
.main-container {
  max-width: 1200px;
  margin: 0 auto;
  padding: 0 32px 64px;
}

/* ══════════════════════════════════════════════════════════════
   STEP WIZARD — 步骤指示器
   ══════════════════════════════════════════════════════════════ */
.steps-bar {
  display: flex;
  align-items: center;
  justify-content: center;
  gap: 0;
  margin: 32px 0 36px;
}
.step-item {
  display: flex;
  align-items: center;
  gap: 10px;
  padding: 10px 20px;
  border-radius: 10px;
  transition: var(--transition);
  cursor: default;
}
.step-item.active { background: var(--accent-soft); }
.step-item.done { background: var(--success-soft); }
.step-num {
  width: 28px; height: 28px;
  border-radius: 50%;
  background: var(--bg-subtle);
  color: var(--text-tertiary);
  font-size: 13px;
  font-weight: 600;
  display: flex;
  align-items: center;
  justify-content: center;
  transition: var(--transition);
}
.step-item.active .step-num {
  background: var(--accent);
  color: white;
}
.step-item.done .step-num {
  background: var(--success);
  color: white;
}
.step-label {
  font-size: 14px;
  font-weight: 500;
  color: var(--text-tertiary);
  transition: var(--transition);
}
.step-item.active .step-label { color: var(--accent-text); }
.step-item.done .step-label { color: var(--success); }
.step-connector {
  width: 40px;
  height: 2px;
  background: var(--border-light);
  border-radius: 1px;
  transition: var(--transition);
}
.step-connector.done { background: var(--success); }

/* ══════════════════════════════════════════════════════════════
   CARDS
   ══════════════════════════════════════════════════════════════ */
.card {
  background: var(--bg-card);
  border: 1px solid var(--border-light);
  border-radius: var(--radius-lg);
  padding: 32px;
  box-shadow: var(--shadow-sm);
  transition: var(--transition);
  margin-bottom: 24px;
}
.card:hover { box-shadow: var(--shadow-md); }
.card-title {
  font-size: 18px;
  font-weight: 600;
  color: var(--text-primary);
  margin-bottom: 6px;
  display: flex;
  align-items: center;
  gap: 10px;
}
.card-title .icon {
  width: 32px; height: 32px;
  border-radius: 9px;
  display: flex;
  align-items: center;
  justify-content: center;
  font-size: 16px;
  flex-shrink: 0;
}
.card-desc {
  font-size: 14px;
  color: var(--text-tertiary);
  margin-bottom: 24px;
}

/* ══════════════════════════════════════════════════════════════
   SECTION PANELS (for each step)
   ══════════════════════════════════════════════════════════════ */
.step-panel { display: none; }
.step-panel.active { display: block; animation: fadeSlideIn 0.35s ease; }

@keyframes fadeSlideIn {
  from { opacity: 0; transform: translateY(12px); }
  to   { opacity: 1; transform: translateY(0); }
}

/* ══════════════════════════════════════════════════════════════
   FILE UPLOAD ZONE
   ══════════════════════════════════════════════════════════════ */
.upload-zone {
  border: 2px dashed var(--border-medium);
  border-radius: var(--radius-lg);
  padding: 56px 32px;
  text-align: center;
  transition: var(--transition);
  cursor: pointer;
  position: relative;
  background: var(--bg-elevated);
}
.upload-zone:hover, .upload-zone.drag-over {
  border-color: var(--accent);
  background: var(--accent-soft);
}
.upload-zone .upload-icon {
  font-size: 48px;
  margin-bottom: 16px;
  opacity: 0.6;
}
.upload-zone h3 {
  font-size: 17px;
  font-weight: 600;
  margin-bottom: 8px;
}
.upload-zone p {
  font-size: 14px;
  color: var(--text-tertiary);
}
.upload-zone input[type="file"] {
  position: absolute;
  inset: 0;
  opacity: 0;
  cursor: pointer;
}

.file-info-card {
  display: none;
  padding: 20px 24px;
  background: var(--success-soft);
  border: 1px solid #A7F3D0;
  border-radius: var(--radius-md);
  margin-top: 20px;
}
.file-info-card.show { display: flex; align-items: center; gap: 16px; }
.file-icon-box {
  width: 48px; height: 48px;
  background: white;
  border-radius: var(--radius-sm);
  display: flex;
  align-items: center;
  justify-content: center;
  font-size: 24px;
  box-shadow: var(--shadow-sm);
  flex-shrink: 0;
}
.file-meta h4 { font-size: 15px; font-weight: 600; margin-bottom: 2px; }
.file-meta p { font-size: 13px; color: var(--text-secondary); }
.file-remove-btn {
  margin-left: auto;
  background: none;
  border: none;
  font-size: 20px;
  cursor: pointer;
  color: var(--text-tertiary);
  padding: 4px 8px;
  border-radius: 6px;
  transition: var(--transition);
}
.file-remove-btn:hover { background: var(--error-soft); color: var(--error); }

/* Sheets list */
.sheets-list {
  margin-top: 16px;
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(260px, 1fr));
  gap: 10px;
}
.sheet-chip {
  display: flex;
  align-items: center;
  gap: 10px;
  padding: 10px 14px;
  background: var(--bg-elevated);
  border: 1px solid var(--border-light);
  border-radius: var(--radius-sm);
  font-size: 14px;
  transition: var(--transition);
}
.sheet-chip.hidden-sheet { opacity: 0.55; }
.sheet-chip label { display: flex; align-items: center; gap: 8px; cursor: pointer; flex:1; }
.sheet-chip input[type="checkbox"] {
  width: 16px; height: 16px;
  accent-color: var(--accent);
  cursor: pointer;
}
.sheet-meta { font-size: 12px; color: var(--text-tertiary); margin-left: auto; }

/* ══════════════════════════════════════════════════════════════
   FORM CONTROLS
   ══════════════════════════════════════════════════════════════ */
.form-grid {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 20px;
}
.form-group { display: flex; flex-direction: column; gap: 6px; }
.form-group.full-width { grid-column: 1 / -1; }
.form-label {
  font-size: 13px;
  font-weight: 600;
  color: var(--text-secondary);
  letter-spacing: 0.2px;
}
.form-hint {
  font-size: 12px;
  color: var(--text-tertiary);
}
.form-select, .form-input {
  padding: 10px 14px;
  border: 1px solid var(--border-light);
  border-radius: var(--radius-sm);
  font-size: 14px;
  font-family: var(--font-sans);
  background: var(--bg-card);
  color: var(--text-primary);
  transition: var(--transition);
  outline: none;
}
.form-select:focus, .form-input:focus {
  border-color: var(--accent);
  box-shadow: 0 0 0 3px rgba(37,99,235,0.1);
}
.form-input::placeholder { color: var(--text-tertiary); }

/* Toggle switches */
.toggle-group {
  display: flex;
  align-items: center;
  justify-content: space-between;
  padding: 12px 0;
  border-bottom: 1px solid var(--bg-elevated);
}
.toggle-group:last-child { border-bottom: none; }
.toggle-info { flex: 1; }
.toggle-info .toggle-name { font-size: 14px; font-weight: 500; }
.toggle-info .toggle-desc { font-size: 12px; color: var(--text-tertiary); margin-top: 2px; }

.toggle-switch {
  position: relative;
  width: 44px; height: 24px;
  flex-shrink: 0;
}
.toggle-switch input { opacity: 0; width: 0; height: 0; }
.toggle-slider {
  position: absolute;
  inset: 0;
  background: var(--border-medium);
  border-radius: 12px;
  cursor: pointer;
  transition: var(--transition);
}
.toggle-slider::before {
  content: '';
  position: absolute;
  width: 18px; height: 18px;
  left: 3px; top: 3px;
  background: white;
  border-radius: 50%;
  transition: var(--transition);
  box-shadow: 0 1px 3px rgba(0,0,0,0.15);
}
.toggle-switch input:checked + .toggle-slider { background: var(--accent); }
.toggle-switch input:checked + .toggle-slider::before { transform: translateX(20px); }

/* Engine cards */
.engine-grid {
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(210px, 1fr));
  gap: 12px;
  margin-bottom: 24px;
}
.engine-card {
  padding: 18px;
  border: 2px solid var(--border-light);
  border-radius: var(--radius-md);
  cursor: pointer;
  transition: var(--transition);
  position: relative;
  background: var(--bg-card);
}
.engine-card:hover { border-color: var(--accent); box-shadow: var(--shadow-md); }
.engine-card.selected {
  border-color: var(--accent);
  background: var(--accent-soft);
  box-shadow: 0 0 0 3px rgba(37,99,235,0.12);
}
.engine-card .engine-name {
  font-size: 15px;
  font-weight: 600;
  margin-bottom: 4px;
}
.engine-card .engine-desc {
  font-size: 12px;
  color: var(--text-tertiary);
}
.engine-card .engine-badge {
  position: absolute;
  top: 10px; right: 10px;
  font-size: 10px;
  font-weight: 600;
  padding: 2px 7px;
  border-radius: 4px;
}
.badge-free { background: var(--success-soft); color: var(--success); }
.badge-key { background: var(--warning-soft); color: var(--warning); }
.badge-env { background: var(--info-soft); color: var(--info); }

/* ══════════════════════════════════════════════════════════════
   BUTTONS
   ══════════════════════════════════════════════════════════════ */
.btn {
  display: inline-flex;
  align-items: center;
  justify-content: center;
  gap: 8px;
  padding: 11px 24px;
  border: none;
  border-radius: var(--radius-sm);
  font-size: 14px;
  font-weight: 600;
  font-family: var(--font-sans);
  cursor: pointer;
  transition: var(--transition);
  text-decoration: none;
}
.btn:disabled { opacity: 0.5; cursor: not-allowed; }
.btn-primary {
  background: var(--accent);
  color: white;
  box-shadow: 0 1px 3px rgba(37,99,235,0.3);
}
.btn-primary:hover:not(:disabled) {
  background: var(--accent-hover);
  box-shadow: 0 4px 12px rgba(37,99,235,0.3);
  transform: translateY(-1px);
}
.btn-secondary {
  background: var(--bg-elevated);
  color: var(--text-secondary);
  border: 1px solid var(--border-light);
}
.btn-secondary:hover:not(:disabled) {
  background: var(--bg-subtle);
  color: var(--text-primary);
}
.btn-success {
  background: var(--success);
  color: white;
}
.btn-success:hover:not(:disabled) { background: #047857; transform: translateY(-1px); }
.btn-ghost {
  background: transparent;
  color: var(--text-secondary);
}
.btn-ghost:hover { background: var(--bg-elevated); }

.btn-row {
  display: flex;
  gap: 12px;
  justify-content: flex-end;
  margin-top: 28px;
}

/* ══════════════════════════════════════════════════════════════
   PROGRESS PANEL
   ══════════════════════════════════════════════════════════════ */
.progress-wrapper {
  background: var(--bg-card);
  border: 1px solid var(--border-light);
  border-radius: var(--radius-lg);
  padding: 36px;
  box-shadow: var(--shadow-md);
  margin-bottom: 24px;
}
.progress-header {
  display: flex;
  align-items: center;
  justify-content: space-between;
  margin-bottom: 20px;
}
.progress-header h3 { font-size: 18px; font-weight: 600; }
.progress-pct {
  font-size: 32px;
  font-weight: 700;
  font-family: var(--font-mono);
  color: var(--accent);
}
.progress-bar-outer {
  width: 100%;
  height: 12px;
  background: var(--bg-subtle);
  border-radius: 6px;
  overflow: hidden;
  margin-bottom: 20px;
}
.progress-bar-inner {
  height: 100%;
  width: 0%;
  background: linear-gradient(90deg, var(--accent), #7C3AED);
  border-radius: 6px;
  transition: width 0.4s ease;
  position: relative;
}
.progress-bar-inner::after {
  content: '';
  position: absolute;
  inset: 0;
  background: linear-gradient(90deg, transparent 0%, rgba(255,255,255,0.3) 50%, transparent 100%);
  animation: shimmer 2s infinite;
}
@keyframes shimmer {
  from { transform: translateX(-100%); }
  to   { transform: translateX(100%); }
}

.progress-stats {
  display: grid;
  grid-template-columns: repeat(4, 1fr);
  gap: 16px;
}
.stat-box {
  text-align: center;
  padding: 14px;
  background: var(--bg-elevated);
  border-radius: var(--radius-sm);
}
.stat-box .stat-value {
  font-size: 22px;
  font-weight: 700;
  font-family: var(--font-mono);
  color: var(--text-primary);
}
.stat-box .stat-label {
  font-size: 12px;
  color: var(--text-tertiary);
  margin-top: 4px;
}

.progress-log {
  margin-top: 20px;
  padding: 16px;
  background: var(--bg-elevated);
  border-radius: var(--radius-sm);
  max-height: 160px;
  overflow-y: auto;
  font-family: var(--font-mono);
  font-size: 13px;
  color: var(--text-secondary);
  line-height: 1.8;
}
.log-entry { padding: 2px 0; }
.log-entry .log-time { color: var(--text-tertiary); }

/* ══════════════════════════════════════════════════════════════
   REPORT / RESULTS PANEL
   ══════════════════════════════════════════════════════════════ */
.result-hero {
  text-align: center;
  padding: 40px 20px;
  margin-bottom: 28px;
}
.result-hero .check-icon {
  width: 72px; height: 72px;
  border-radius: 50%;
  background: var(--success-soft);
  display: flex;
  align-items: center;
  justify-content: center;
  margin: 0 auto 16px;
  font-size: 36px;
  color: var(--success);
  animation: popIn 0.4s cubic-bezier(0.34, 1.56, 0.64, 1);
}
@keyframes popIn {
  from { transform: scale(0); } to { transform: scale(1); }
}
.result-hero h2 { font-size: 26px; font-weight: 700; margin-bottom: 8px; }
.result-hero p { color: var(--text-secondary); font-size: 15px; }

.download-cards {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
  gap: 16px;
  margin-bottom: 32px;
}
.download-card {
  display: flex;
  align-items: center;
  gap: 16px;
  padding: 20px 24px;
  border: 1px solid var(--border-light);
  border-radius: var(--radius-md);
  background: var(--bg-card);
  transition: var(--transition);
}
.download-card:hover { box-shadow: var(--shadow-md); border-color: var(--accent); }
.dl-icon {
  width: 44px; height: 44px;
  border-radius: 10px;
  display: flex;
  align-items: center;
  justify-content: center;
  font-size: 22px;
  flex-shrink: 0;
}
.dl-info { flex: 1; }
.dl-info h4 { font-size: 14px; font-weight: 600; margin-bottom: 2px; }
.dl-info p { font-size: 12px; color: var(--text-tertiary); }

/* Report table */
.report-table {
  width: 100%;
  border-collapse: separate;
  border-spacing: 0;
  font-size: 13px;
}
.report-table th {
  padding: 10px 14px;
  text-align: left;
  font-weight: 600;
  color: var(--text-secondary);
  background: var(--bg-elevated);
  border-bottom: 1px solid var(--border-light);
}
.report-table th:first-child { border-radius: var(--radius-sm) 0 0 0; }
.report-table th:last-child { border-radius: 0 var(--radius-sm) 0 0; }
.report-table td {
  padding: 10px 14px;
  border-bottom: 1px solid var(--bg-elevated);
  font-family: var(--font-mono);
  font-size: 13px;
}
.report-table tr:last-child td { border-bottom: none; }
.report-table .total-row {
  font-weight: 700;
  background: var(--bg-elevated);
}

.report-meta-grid {
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(180px, 1fr));
  gap: 12px;
  margin-bottom: 24px;
}
.meta-item {
  padding: 14px;
  background: var(--bg-elevated);
  border-radius: var(--radius-sm);
}
.meta-item .meta-label { font-size: 12px; color: var(--text-tertiary); margin-bottom: 4px; }
.meta-item .meta-value {
  font-size: 18px;
  font-weight: 700;
  font-family: var(--font-mono);
}

/* ══════════════════════════════════════════════════════════════
   COLLAPSIBLE ADVANCED SETTINGS
   ══════════════════════════════════════════════════════════════ */
.collapsible-header {
  display: flex;
  align-items: center;
  gap: 8px;
  padding: 12px 0;
  cursor: pointer;
  font-size: 14px;
  font-weight: 600;
  color: var(--text-secondary);
  user-select: none;
  transition: var(--transition);
}
.collapsible-header:hover { color: var(--accent); }
.collapsible-header .arrow {
  transition: transform 0.25s ease;
  font-size: 12px;
}
.collapsible-header.open .arrow { transform: rotate(90deg); }
.collapsible-body {
  max-height: 0;
  overflow: hidden;
  transition: max-height 0.35s ease, padding 0.35s ease;
}
.collapsible-body.open {
  max-height: 1200px;
  padding-bottom: 16px;
}

/* ══════════════════════════════════════════════════════════════
   ERROR TOAST
   ══════════════════════════════════════════════════════════════ */
.toast-container {
  position: fixed;
  top: 80px;
  right: 24px;
  z-index: 9999;
  display: flex;
  flex-direction: column;
  gap: 8px;
}
.toast {
  padding: 14px 20px;
  border-radius: var(--radius-sm);
  font-size: 14px;
  font-weight: 500;
  box-shadow: var(--shadow-lg);
  animation: slideInRight 0.3s ease;
  max-width: 420px;
}
@keyframes slideInRight {
  from { transform: translateX(100%); opacity: 0; }
  to   { transform: translateX(0); opacity: 1; }
}
.toast-error { background: var(--error); color: white; }
.toast-success { background: var(--success); color: white; }
.toast-info { background: var(--accent); color: white; }

/* ══════════════════════════════════════════════════════════════
   RESPONSIVE
   ══════════════════════════════════════════════════════════════ */
@media (max-width: 768px) {
  .hero h1 { font-size: 28px; }
  .form-grid { grid-template-columns: 1fr; }
  .progress-stats { grid-template-columns: repeat(2, 1fr); }
  .engine-grid { grid-template-columns: 1fr 1fr; }
  .steps-bar { flex-wrap: wrap; }
  .step-connector { display: none; }
  .header-inner { padding: 0 16px; }
  .main-container { padding: 0 16px 40px; }
  .card { padding: 20px; }
}

/* ══════════════════════════════════════════════════════════════
   FOOTER
   ══════════════════════════════════════════════════════════════ */
.site-footer {
  text-align: center;
  padding: 32px;
  font-size: 13px;
  color: var(--text-tertiary);
  border-top: 1px solid var(--border-light);
  margin-top: 40px;
}

/* Loading spinner */
.spinner {
  width: 20px; height: 20px;
  border: 3px solid rgba(255,255,255,0.3);
  border-top-color: white;
  border-radius: 50%;
  animation: spin 0.7s linear infinite;
  display: inline-block;
}
@keyframes spin { to { transform: rotate(360deg); } }
</style>
</head>
<body>

<!-- ════════════════════════════════════════════════════════════
     HEADER
     ════════════════════════════════════════════════════════════ -->
<header class="site-header">
  <div class="header-inner">
    <a href="/" class="logo">
      <div class="logo-icon">E</div>
      <div class="logo-text">Excel<span>Translator</span> Pro</div>
    </a>
    <span class="header-badge">v""" + VERSION + r"""</span>
  </div>
</header>

<!-- ════════════════════════════════════════════════════════════
     HERO
     ════════════════════════════════════════════════════════════ -->
<section class="hero">
  <h1>智能 Excel 多语言翻译</h1>
  <p>上传 Excel 文件, 选择翻译引擎, 一键完成全文翻译。<br>完美保留所有格式 · 支持 5 大翻译引擎 · 实时翻译进度</p>
  <div class="engine-pills">
    <span class="engine-pill">DeepSeek</span>
    <span class="engine-pill">OpenAI GPT</span>
    <span class="engine-pill">Claude</span>
    <span class="engine-pill">Google 翻译</span>
    <span class="engine-pill">通义千问</span>
  </div>
</section>

<!-- ════════════════════════════════════════════════════════════
     STEP WIZARD BAR
     ════════════════════════════════════════════════════════════ -->
<div class="main-container">
  <div class="steps-bar" id="stepsBar">
    <div class="step-item active" data-step="1">
      <div class="step-num">1</div>
      <span class="step-label">上传文件</span>
    </div>
    <div class="step-connector" id="conn1"></div>
    <div class="step-item" data-step="2">
      <div class="step-num">2</div>
      <span class="step-label">翻译配置</span>
    </div>
    <div class="step-connector" id="conn2"></div>
    <div class="step-item" data-step="3">
      <div class="step-num">3</div>
      <span class="step-label">执行翻译</span>
    </div>
    <div class="step-connector" id="conn3"></div>
    <div class="step-item" data-step="4">
      <div class="step-num">4</div>
      <span class="step-label">下载结果</span>
    </div>
  </div>

  <!-- ══════════════════════════════════════════════════════════
       STEP 1: 上传文件
       ══════════════════════════════════════════════════════════ -->
  <div class="step-panel active" id="step1">
    <div class="card">
      <div class="card-title">
        <div class="icon" style="background:var(--accent-soft);color:var(--accent);">📁</div>
        上传 Excel 文件
      </div>
      <div class="card-desc">支持 .xlsx 和 .xlsm 格式，拖拽或点击上传</div>

      <div class="upload-zone" id="uploadZone">
        <input type="file" id="fileInput" accept=".xlsx,.xlsm">
        <div class="upload-icon">📊</div>
        <h3>拖拽文件到此处 或 点击选择文件</h3>
        <p>支持 .xlsx / .xlsm 格式</p>
      </div>

      <div class="file-info-card" id="fileInfoCard">
        <div class="file-icon-box">📄</div>
        <div class="file-meta">
          <h4 id="fileName">-</h4>
          <p id="fileMeta">-</p>
        </div>
        <button class="file-remove-btn" onclick="removeFile()" title="移除文件">✕</button>
      </div>

      <div id="sheetsArea" style="display:none; margin-top:24px;">
        <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px;">
          <span style="font-size:14px;font-weight:600;color:var(--text-secondary);">工作表列表</span>
          <label style="font-size:13px;color:var(--text-tertiary);cursor:pointer;display:flex;align-items:center;gap:6px;">
            <input type="checkbox" id="selectAllSheets" checked onchange="toggleAllSheets(this.checked)" style="accent-color:var(--accent);">
            全选
          </label>
        </div>
        <div class="sheets-list" id="sheetsList"></div>
      </div>

      <div class="btn-row">
        <button class="btn btn-primary" id="btnToStep2" disabled onclick="goToStep(2)">
          下一步：翻译配置 →
        </button>
      </div>
    </div>
  </div>

  <!-- ══════════════════════════════════════════════════════════
       STEP 2: 翻译配置
       ══════════════════════════════════════════════════════════ -->
  <div class="step-panel" id="step2">
    <div class="card">
      <div class="card-title">
        <div class="icon" style="background:var(--info-soft);color:var(--info);">🔧</div>
        翻译引擎
      </div>
      <div class="card-desc">选择您偏好的翻译引擎</div>

      <div class="engine-grid" id="engineGrid">
        <div class="engine-card selected" data-engine="deepseek" onclick="selectEngine('deepseek')">
          <div class="engine-name">DeepSeek</div>
          <div class="engine-desc">高质量翻译 · 成本低</div>
          <span class="engine-badge badge-key">API Key</span>
        </div>
        <div class="engine-card" data-engine="openai" onclick="selectEngine('openai')">
          <div class="engine-name">OpenAI GPT</div>
          <div class="engine-desc">GPT-4o-mini · 通用翻译</div>
          <span class="engine-badge badge-key">API Key</span>
        </div>
        <div class="engine-card" data-engine="claude" onclick="selectEngine('claude')">
          <div class="engine-name">Claude</div>
          <div class="engine-desc">高质量 · 擅长专业术语</div>
          <span class="engine-badge badge-key">API Key</span>
        </div>
        <div class="engine-card" data-engine="google_free" onclick="selectEngine('google_free')">
          <div class="engine-name">Google 翻译</div>
          <div class="engine-desc">免费引擎 · 无需密钥</div>
          <span class="engine-badge badge-free">免费</span>
        </div>
        <div class="engine-card" data-engine="qwen" onclick="selectEngine('qwen')">
          <div class="engine-name">通义千问</div>
          <div class="engine-desc">阿里云 · 中文翻译优秀</div>
          <span class="engine-badge badge-key">API Key</span>
        </div>
      </div>

      <!-- API Key -->
      <div class="form-group" id="apiKeyGroup">
        <label class="form-label">API Key</label>
        <input class="form-input" type="password" id="apiKeyInput" placeholder="输入 API Key (或已在环境变量中配置)">
        <div class="form-hint" id="apiKeyHint">环境变量: DEEPSEEK_API_KEY</div>
      </div>
    </div>

    <div class="card">
      <div class="card-title">
        <div class="icon" style="background:var(--success-soft);color:var(--success);">🌐</div>
        语言设置
      </div>

      <div class="form-grid">
        <div class="form-group">
          <label class="form-label">源语言</label>
          <select class="form-select" id="sourceLang"></select>
        </div>
        <div class="form-group">
          <label class="form-label">目标语言</label>
          <select class="form-select" id="targetLang"></select>
        </div>
      </div>
    </div>

    <!-- Advanced settings -->
    <div class="card">
      <div class="collapsible-header" onclick="toggleCollapsible(this)">
        <span class="arrow">▶</span>
        ⚙️ 高级设置
      </div>
      <div class="collapsible-body" id="advancedBody">
        <div class="form-grid" style="margin-bottom:20px;">
          <div class="form-group">
            <label class="form-label">自定义模型名称</label>
            <input class="form-input" id="modelInput" placeholder="留空使用引擎默认模型">
          </div>
          <div class="form-group">
            <label class="form-label">自定义 API Base URL</label>
            <input class="form-input" id="apiBaseInput" placeholder="留空使用默认地址 (代理/私有部署)">
          </div>
          <div class="form-group">
            <label class="form-label">批量翻译大小</label>
            <input class="form-input" type="number" id="batchSize" value="30" min="1" max="100">
          </div>
          <div class="form-group">
            <label class="form-label">并发线程数</label>
            <input class="form-input" type="number" id="maxWorkers" value="3" min="1" max="10">
          </div>
          <div class="form-group">
            <label class="form-label">最大重试次数</label>
            <input class="form-input" type="number" id="maxRetries" value="3" min="1" max="10">
          </div>
        </div>

        <div style="border-top:1px solid var(--bg-elevated); padding-top:16px;">
          <div class="toggle-group">
            <div class="toggle-info">
              <div class="toggle-name">生成双语对照版本</div>
              <div class="toggle-desc">在原文右侧插入翻译列，方便逐一对照审核</div>
            </div>
            <label class="toggle-switch">
              <input type="checkbox" id="bilingualToggle">
              <span class="toggle-slider"></span>
            </label>
          </div>
          <div class="toggle-group">
            <div class="toggle-info">
              <div class="toggle-name">翻译单元格批注</div>
              <div class="toggle-desc">同时翻译单元格中的批注 (Comment) 内容</div>
            </div>
            <label class="toggle-switch">
              <input type="checkbox" id="translateCommentsToggle">
              <span class="toggle-slider"></span>
            </label>
          </div>
          <div class="toggle-group">
            <div class="toggle-info">
              <div class="toggle-name">高亮标记翻译失败单元格</div>
              <div class="toggle-desc">用黄色背景标记翻译失败的单元格</div>
            </div>
            <label class="toggle-switch">
              <input type="checkbox" id="highlightToggle" checked>
              <span class="toggle-slider"></span>
            </label>
          </div>
          <div class="toggle-group">
            <div class="toggle-info">
              <div class="toggle-name">跳过已是目标语言的文本</div>
              <div class="toggle-desc">检测已翻译内容，避免重复翻译</div>
            </div>
            <label class="toggle-switch">
              <input type="checkbox" id="skipTargetToggle" checked>
              <span class="toggle-slider"></span>
            </label>
          </div>
          <div class="toggle-group">
            <div class="toggle-info">
              <div class="toggle-name">跳过隐藏工作表</div>
              <div class="toggle-desc">不翻译隐藏状态的 Sheet</div>
            </div>
            <label class="toggle-switch">
              <input type="checkbox" id="skipHiddenToggle" checked>
              <span class="toggle-slider"></span>
            </label>
          </div>
          <div class="toggle-group">
            <div class="toggle-info">
              <div class="toggle-name">启用翻译缓存</div>
              <div class="toggle-desc">缓存已翻译文本，支持断点续传</div>
            </div>
            <label class="toggle-switch">
              <input type="checkbox" id="cacheToggle" checked>
              <span class="toggle-slider"></span>
            </label>
          </div>
        </div>
      </div>
    </div>

    <div class="btn-row">
      <button class="btn btn-secondary" onclick="goToStep(1)">← 返回</button>
      <button class="btn btn-secondary" id="btnDryRun" onclick="startTask(true)">
        📊 仅分析 (Dry-run)
      </button>
      <button class="btn btn-primary" id="btnStartTranslate" onclick="startTask(false)">
        🚀 开始翻译
      </button>
    </div>
  </div>

  <!-- ══════════════════════════════════════════════════════════
       STEP 3: 执行翻译 (进度)
       ══════════════════════════════════════════════════════════ -->
  <div class="step-panel" id="step3">
    <div class="progress-wrapper">
      <div class="progress-header">
        <h3 id="progressTitle">正在翻译中...</h3>
        <div class="progress-pct" id="progressPct">0%</div>
      </div>
      <div class="progress-bar-outer">
        <div class="progress-bar-inner" id="progressBar"></div>
      </div>
      <div class="progress-stats">
        <div class="stat-box">
          <div class="stat-value" id="statTotal">0</div>
          <div class="stat-label">总计</div>
        </div>
        <div class="stat-box">
          <div class="stat-value" id="statDone" style="color:var(--success);">0</div>
          <div class="stat-label">已完成</div>
        </div>
        <div class="stat-box">
          <div class="stat-value" id="statCached" style="color:var(--info);">0</div>
          <div class="stat-label">缓存命中</div>
        </div>
        <div class="stat-box">
          <div class="stat-value" id="statSheet">-</div>
          <div class="stat-label">当前 Sheet</div>
        </div>
      </div>
      <div class="progress-log" id="progressLog"></div>
    </div>
  </div>

  <!-- ══════════════════════════════════════════════════════════
       STEP 4: 下载结果
       ══════════════════════════════════════════════════════════ -->
  <div class="step-panel" id="step4">
    <div class="result-hero" id="resultHero">
      <div class="check-icon">✓</div>
      <h2 id="resultTitle">翻译完成</h2>
      <p id="resultSubtitle">所有单元格已成功翻译，文件已准备好下载</p>
    </div>

    <div class="download-cards" id="downloadCards"></div>

    <div class="card">
      <div class="card-title">
        <div class="icon" style="background:var(--warning-soft);color:var(--warning);">📊</div>
        翻译报告
      </div>

      <div class="report-meta-grid" id="reportMeta"></div>

      <div style="overflow-x:auto;">
        <table class="report-table" id="reportTable">
          <thead>
            <tr>
              <th>Sheet 名称</th>
              <th>需翻译</th>
              <th>成功</th>
              <th>缓存</th>
              <th>失败</th>
              <th>跳过</th>
            </tr>
          </thead>
          <tbody id="reportBody"></tbody>
        </table>
      </div>
    </div>

    <div class="btn-row">
      <button class="btn btn-secondary" onclick="resetAll()">🔄 翻译新文件</button>
    </div>
  </div>

</div>

<!-- Toast container -->
<div class="toast-container" id="toastContainer"></div>

<!-- Footer -->
<footer class="site-footer">
  ExcelTranslator Pro v""" + VERSION + r""" · Powered by DeepSeek / OpenAI / Claude / Google / Qwen
</footer>

<!-- ════════════════════════════════════════════════════════════
     JAVASCRIPT
     ════════════════════════════════════════════════════════════ -->
<script>
// ─── 全局状态 ───
let fileData = null;     // { file_id, filename, filepath, sheets }
let currentStep = 1;
let taskId = null;
let selectedEngine = 'deepseek';
let eventSource = null;

const ENV_VARS = {
  deepseek: 'DEEPSEEK_API_KEY',
  openai: 'OPENAI_API_KEY',
  claude: 'ANTHROPIC_API_KEY',
  google_free: null,
  qwen: 'DASHSCOPE_API_KEY',
};

const LANG_MAP = """ + json.dumps(LANGUAGE_NAMES, ensure_ascii=False) + r""";

// ─── 初始化 ───
document.addEventListener('DOMContentLoaded', () => {
  initLanguageSelects();
  initDragDrop();
  loadEngineStatus();
});

function initLanguageSelects() {
  const src = document.getElementById('sourceLang');
  const tgt = document.getElementById('targetLang');
  for (const [code, name] of Object.entries(LANG_MAP)) {
    src.innerHTML += `<option value="${code}" ${code === 'ru' ? 'selected' : ''}>${name}</option>`;
    tgt.innerHTML += `<option value="${code}" ${code === 'zh-CN' ? 'selected' : ''}>${name}</option>`;
  }
}

function initDragDrop() {
  const zone = document.getElementById('uploadZone');
  zone.addEventListener('dragover', e => { e.preventDefault(); zone.classList.add('drag-over'); });
  zone.addEventListener('dragleave', () => zone.classList.remove('drag-over'));
  zone.addEventListener('drop', e => {
    e.preventDefault();
    zone.classList.remove('drag-over');
    if (e.dataTransfer.files.length) {
      document.getElementById('fileInput').files = e.dataTransfer.files;
      handleFileSelect(e.dataTransfer.files[0]);
    }
  });
  document.getElementById('fileInput').addEventListener('change', e => {
    if (e.target.files.length) handleFileSelect(e.target.files[0]);
  });
}

async function loadEngineStatus() {
  try {
    const resp = await fetch('/api/engines');
    const data = await resp.json();
    // Update engine badges based on env status
    for (const [eng, hasKey] of Object.entries(data.env_status)) {
      const card = document.querySelector(`.engine-card[data-engine="${eng}"]`);
      if (card && hasKey && eng !== 'google_free') {
        const badge = card.querySelector('.engine-badge');
        if (badge) {
          badge.className = 'engine-badge badge-env';
          badge.textContent = 'ENV ✓';
        }
      }
    }
  } catch (e) { /* ignore */ }
}

// ─── 文件上传 ───
async function handleFileSelect(file) {
  if (!file) return;
  const ext = file.name.split('.').pop().toLowerCase();
  if (!['xlsx', 'xlsm'].includes(ext)) {
    showToast('仅支持 .xlsx / .xlsm 格式', 'error');
    return;
  }

  const formData = new FormData();
  formData.append('file', file);

  try {
    document.getElementById('uploadZone').style.display = 'none';
    document.getElementById('fileInfoCard').classList.add('show');
    document.getElementById('fileName').textContent = file.name;
    document.getElementById('fileMeta').textContent = '正在解析...';

    const resp = await fetch('/api/upload', { method: 'POST', body: formData });
    const data = await resp.json();

    if (data.error) {
      showToast(data.error, 'error');
      removeFile();
      return;
    }

    fileData = data;
    document.getElementById('fileMeta').textContent =
      `${data.sheets.length} 个工作表 · ${(file.size / 1024).toFixed(1)} KB`;

    // Render sheets
    const listEl = document.getElementById('sheetsList');
    listEl.innerHTML = '';
    for (const s of data.sheets) {
      const hidden = s.hidden ? ' hidden-sheet' : '';
      listEl.innerHTML += `
        <div class="sheet-chip${hidden}">
          <label>
            <input type="checkbox" class="sheet-cb" value="${s.name}" ${s.hidden ? '' : 'checked'}>
            ${s.name}${s.hidden ? ' <small style="color:var(--text-tertiary)">(隐藏)</small>' : ''}
          </label>
          <span class="sheet-meta">${s.rows}×${s.cols}</span>
        </div>
      `;
    }
    document.getElementById('sheetsArea').style.display = 'block';
    document.getElementById('btnToStep2').disabled = false;
  } catch (e) {
    showToast('文件上传失败: ' + e.message, 'error');
    removeFile();
  }
}

function removeFile() {
  fileData = null;
  document.getElementById('uploadZone').style.display = '';
  document.getElementById('fileInfoCard').classList.remove('show');
  document.getElementById('sheetsArea').style.display = 'none';
  document.getElementById('fileInput').value = '';
  document.getElementById('btnToStep2').disabled = true;
}

function toggleAllSheets(checked) {
  document.querySelectorAll('.sheet-cb').forEach(cb => cb.checked = checked);
}

// ─── 引擎选择 ───
function selectEngine(eng) {
  selectedEngine = eng;
  document.querySelectorAll('.engine-card').forEach(c => c.classList.remove('selected'));
  document.querySelector(`.engine-card[data-engine="${eng}"]`).classList.add('selected');

  const keyGroup = document.getElementById('apiKeyGroup');
  const hint = document.getElementById('apiKeyHint');
  if (eng === 'google_free') {
    keyGroup.style.display = 'none';
  } else {
    keyGroup.style.display = '';
    hint.textContent = `环境变量: ${ENV_VARS[eng] || ''}`;
  }
}

// ─── 折叠面板 ───
function toggleCollapsible(header) {
  header.classList.toggle('open');
  const body = header.nextElementSibling;
  body.classList.toggle('open');
}

// ─── 步骤导航 ───
function goToStep(step) {
  document.querySelectorAll('.step-panel').forEach(p => p.classList.remove('active'));
  document.getElementById(`step${step}`).classList.add('active');

  document.querySelectorAll('.step-item').forEach(s => {
    const sn = parseInt(s.dataset.step);
    s.classList.remove('active', 'done');
    if (sn < step) s.classList.add('done');
    else if (sn === step) s.classList.add('active');
  });
  document.querySelectorAll('.step-connector').forEach((c, i) => {
    c.classList.toggle('done', i + 1 < step);
  });
  currentStep = step;
  window.scrollTo({ top: 0, behavior: 'smooth' });
}

// ─── 启动翻译 ───
async function startTask(dryRun) {
  if (!fileData) { showToast('请先上传文件', 'error'); return; }

  // Collect selected sheets
  const selectedSheets = Array.from(document.querySelectorAll('.sheet-cb:checked')).map(c => c.value);
  if (selectedSheets.length === 0) {
    showToast('请至少选择一个工作表', 'error');
    return;
  }

  // Validate API key for non-free engines
  const apiKey = document.getElementById('apiKeyInput').value.trim();
  if (!dryRun && selectedEngine !== 'google_free' && !apiKey) {
    // Check if we have env key info (we don't know for sure client-side, so let server validate)
  }

  const payload = {
    filepath: fileData.filepath,
    filename: fileData.filename,
    engine: selectedEngine,
    api_key: apiKey,
    api_base: document.getElementById('apiBaseInput').value.trim(),
    model: document.getElementById('modelInput').value.trim(),
    source_lang: document.getElementById('sourceLang').value,
    target_lang: document.getElementById('targetLang').value,
    sheets: selectedSheets.join(','),
    batch_size: parseInt(document.getElementById('batchSize').value) || 30,
    max_workers: parseInt(document.getElementById('maxWorkers').value) || 3,
    max_retries: parseInt(document.getElementById('maxRetries').value) || 3,
    bilingual: document.getElementById('bilingualToggle').checked,
    translate_comments: document.getElementById('translateCommentsToggle').checked,
    highlight_failures: document.getElementById('highlightToggle').checked,
    skip_target_lang: document.getElementById('skipTargetToggle').checked,
    skip_hidden: document.getElementById('skipHiddenToggle').checked,
    no_cache: !document.getElementById('cacheToggle').checked,
    dry_run: dryRun,
  };

  // Disable buttons
  document.getElementById('btnStartTranslate').disabled = true;
  document.getElementById('btnDryRun').disabled = true;

  try {
    const resp = await fetch('/api/translate', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify(payload),
    });
    const data = await resp.json();

    if (data.error) {
      showToast(data.error, 'error');
      document.getElementById('btnStartTranslate').disabled = false;
      document.getElementById('btnDryRun').disabled = false;
      return;
    }

    taskId = data.task_id;
    goToStep(3);

    // Reset progress UI
    document.getElementById('progressPct').textContent = '0%';
    document.getElementById('progressBar').style.width = '0%';
    document.getElementById('progressLog').innerHTML = '';
    document.getElementById('statTotal').textContent = '0';
    document.getElementById('statDone').textContent = '0';
    document.getElementById('statCached').textContent = '0';
    document.getElementById('statSheet').textContent = '-';
    document.getElementById('progressTitle').textContent =
      dryRun ? '正在分析文件...' : '正在翻译中...';

    addLog('任务已创建，等待服务器响应...');

    // Start SSE
    startEventStream(taskId);
  } catch (e) {
    showToast('启动失败: ' + e.message, 'error');
    document.getElementById('btnStartTranslate').disabled = false;
    document.getElementById('btnDryRun').disabled = false;
  }
}

// ─── SSE 事件流 ───
function startEventStream(tid) {
  if (eventSource) eventSource.close();
  eventSource = new EventSource(`/api/task/${tid}/stream`);

  eventSource.onmessage = function(e) {
    const evt = JSON.parse(e.data);
    if (!evt || !evt.type) return;

    switch (evt.type) {
      case 'status':
        addLog(evt.data.message);
        break;

      case 'analysis':
        document.getElementById('statTotal').textContent = evt.data.total_translatable;
        addLog(`文件分析完成: 共 ${evt.data.total_translatable} 个单元格需翻译`);
        for (const s of evt.data.sheets) {
          addLog(`  📋 ${s.name}: ${s.translatable} 待翻译, ${s.total_skipped} 跳过`);
        }
        break;

      case 'cache':
        document.getElementById('statCached').textContent = evt.data.cached;
        addLog(`缓存命中 ${evt.data.cached} 个, 剩余 ${evt.data.remaining} 待翻译`);
        break;

      case 'progress':
        const p = evt.data.progress;
        document.getElementById('progressPct').textContent = p + '%';
        document.getElementById('progressBar').style.width = p + '%';
        document.getElementById('statDone').textContent = evt.data.completed;
        document.getElementById('statTotal').textContent = evt.data.total;
        document.getElementById('statSheet').textContent = evt.data.sheet || '-';
        break;

      case 'done':
        eventSource.close();
        addLog('✅ ' + evt.data.message);
        document.getElementById('progressPct').textContent = '100%';
        document.getElementById('progressBar').style.width = '100%';
        setTimeout(() => showResults(evt.data.report), 600);
        break;

      case 'error':
        eventSource.close();
        addLog('❌ 错误: ' + evt.data.error);
        showToast('翻译失败: ' + evt.data.error, 'error');
        document.getElementById('btnStartTranslate').disabled = false;
        document.getElementById('btnDryRun').disabled = false;
        break;

      case 'heartbeat':
        break;
    }
  };

  eventSource.onerror = function() {
    // Don't close - SSE reconnects automatically
  };
}

function addLog(text) {
  const log = document.getElementById('progressLog');
  const now = new Date().toLocaleTimeString('zh-CN', { hour12: false });
  log.innerHTML += `<div class="log-entry"><span class="log-time">[${now}]</span> ${text}</div>`;
  log.scrollTop = log.scrollHeight;
}

// ─── 显示结果 ───
function showResults(report) {
  goToStep(4);

  if (!report) {
    document.getElementById('resultTitle').textContent = '分析完成';
    document.getElementById('resultSubtitle').textContent = 'Dry-run 模式：仅完成文件分析';
    return;
  }

  const isDryRun = report.engine === 'Dry-run';
  if (isDryRun) {
    document.getElementById('resultHero').querySelector('.check-icon').textContent = '📊';
    document.getElementById('resultTitle').textContent = '文件分析完成';
    document.getElementById('resultSubtitle').textContent =
      `共 ${report.total_translatable} 个单元格需翻译，${report.total_skipped} 个跳过`;
  } else {
    document.getElementById('resultTitle').textContent = '翻译完成!';
    document.getElementById('resultSubtitle').textContent =
      `成功翻译 ${report.total_ok} 个单元格${report.total_fail > 0 ? '，' + report.total_fail + ' 个失败' : ''}`;
  }

  // Download cards
  const dlCards = document.getElementById('downloadCards');
  dlCards.innerHTML = '';
  if (!isDryRun) {
    dlCards.innerHTML += `
      <div class="download-card">
        <div class="dl-icon" style="background:var(--success-soft);color:var(--success);">📥</div>
        <div class="dl-info">
          <h4>翻译结果文件</h4>
          <p>保留原始格式的翻译版本</p>
        </div>
        <a class="btn btn-success" href="/api/download/${taskId}/translated" download>下载</a>
      </div>
    `;
    // Check bilingual
    fetch(`/api/task/${taskId}/status`).then(r => r.json()).then(data => {
      if (data.bilingual_file) {
        dlCards.innerHTML += `
          <div class="download-card">
            <div class="dl-icon" style="background:var(--info-soft);color:var(--info);">📑</div>
            <div class="dl-info">
              <h4>双语对照版本</h4>
              <p>原文 + 翻译并排显示</p>
            </div>
            <a class="btn btn-primary" href="/api/download/${taskId}/bilingual" download>下载</a>
          </div>
        `;
      }
    });
  }

  // Report meta
  const metaEl = document.getElementById('reportMeta');
  metaEl.innerHTML = `
    <div class="meta-item"><div class="meta-label">翻译引擎</div><div class="meta-value" style="font-size:14px;">${report.engine || '-'}</div></div>
    <div class="meta-item"><div class="meta-label">总计翻译</div><div class="meta-value" style="color:var(--success);">${report.total_ok}</div></div>
    <div class="meta-item"><div class="meta-label">缓存命中</div><div class="meta-value" style="color:var(--info);">${report.total_cached}</div></div>
    <div class="meta-item"><div class="meta-label">缓存命中率</div><div class="meta-value">${report.cache_hit_rate}%</div></div>
    <div class="meta-item"><div class="meta-label">API 调用</div><div class="meta-value">${report.api_calls}</div></div>
    <div class="meta-item"><div class="meta-label">重试次数</div><div class="meta-value">${report.retries}</div></div>
  `;

  // Report table
  const tbody = document.getElementById('reportBody');
  tbody.innerHTML = '';
  for (const s of report.sheets) {
    tbody.innerHTML += `<tr>
      <td style="font-family:var(--font-sans);font-weight:500;">${s.name}</td>
      <td>${s.translatable}</td>
      <td style="color:var(--success);">${s.translated_ok}</td>
      <td style="color:var(--info);">${s.translated_cached}</td>
      <td style="color:${s.translated_fail > 0 ? 'var(--error)' : 'inherit'};">${s.translated_fail}</td>
      <td>${s.total_skipped}</td>
    </tr>`;
  }
  tbody.innerHTML += `<tr class="total-row">
    <td style="font-family:var(--font-sans);">合计</td>
    <td>${report.total_translatable}</td>
    <td style="color:var(--success);">${report.total_ok}</td>
    <td style="color:var(--info);">${report.total_cached}</td>
    <td style="color:${report.total_fail > 0 ? 'var(--error)' : 'inherit'};">${report.total_fail}</td>
    <td>${report.total_skipped}</td>
  </tr>`;
}

// ─── 重置 ───
function resetAll() {
  fileData = null;
  taskId = null;
  if (eventSource) eventSource.close();
  removeFile();
  document.getElementById('btnStartTranslate').disabled = false;
  document.getElementById('btnDryRun').disabled = false;
  document.getElementById('apiKeyInput').value = '';
  document.getElementById('modelInput').value = '';
  document.getElementById('apiBaseInput').value = '';
  goToStep(1);
}

// ─── Toast 通知 ───
function showToast(msg, type = 'info') {
  const container = document.getElementById('toastContainer');
  const toast = document.createElement('div');
  toast.className = `toast toast-${type}`;
  toast.textContent = msg;
  container.appendChild(toast);
  setTimeout(() => { toast.style.opacity = '0'; setTimeout(() => toast.remove(), 300); }, 5000);
}
</script>
</body>
</html>
"""


# ============================================================================
# 启动服务器
# ============================================================================

def _open_browser():
    """延迟 1.5 秒后自动打开浏览器"""
    time.sleep(1.5)
    webbrowser.open("http://localhost:8686")


if __name__ == "__main__":
    # 初始化日志
    setup_logging(log_level="INFO")

    print("\n" + "=" * 56)
    print(f"  ExcelTranslator Pro  Web v{VERSION}")
    print("=" * 56)
    print(f"  🌐 打开浏览器访问: http://localhost:8686")
    print(f"  📁 上传目录: {UPLOAD_FOLDER}")
    print(f"  📤 输出目录: {OUTPUT_FOLDER}")
    print("=" * 56 + "\n")

    # 启动浏览器自动打开线程
    threading.Thread(target=_open_browser, daemon=True).start()

    app.run(host="0.0.0.0", port=8686, debug=False, threaded=True)